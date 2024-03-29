from .models import *
from ..invitado.managers import *
from ..residente.managers import *
from ..evento.managers import *
from ..areasocial.managers import *
from ..nropase.managers import *
from ..vehiculo.managers import *
from sqlalchemy import or_,and_
from sqlalchemy import and_
from sqlalchemy.exc import IntegrityError
from threading import Thread

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

import threading

class MovimientoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Movimiento, db)

    def reporte_generar(self,diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        # movimiento = self.db.query(self.entity)\
        #     .filter(self.entity.estado == True).all()


        movimiento = self.db.query(self.entity)\
            .filter(func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])) \
            .filter(self.entity.estado == True).all()


        t = Thread(target=self.sincronizar_descripcion, args=(movimiento,))
        t.start()

        return movimiento

    def sincronizar_descripcion_residente(self, objetos):
        print("sincronizar descripcion residente")
        cont = 1

        for objeto in objetos:

            if objeto.fkresidente:
                objeto.descripcion_residente = objeto.residente.fullname
            else:
                objeto.descripcion_residente = '-----'

            self.db.merge(objeto)
            print(str(cont))
            cont = cont + 1

        fecha = BitacoraManager(self.db).fecha_actual()

        print("sincronizar inicio commit " + fecha.strftime('%d/%m/%Y %H:%M:%S'))
        self.db.commit()
        print("sincronizar fin commit " + fecha.strftime('%d/%m/%Y %H:%M:%S'))

    def sincronizar_descripcion(self, objetos):
        print("sincronizar descripcion")
        cont = 1

        for objeto in objetos:

            objeto.descripcion_fechai = objeto.fechar.strftime('%d/%m/%Y %H:%M:%S') if objeto.fechar else '-----'
            objeto.descripcion_fechaf = objeto.fechaf.strftime('%d/%m/%Y %H:%M:%S') if objeto.fechaf else '-----'
            objeto.descripcion_documento = objeto.tipodocumento.nombre

            if objeto.fkresidente:
                objeto.descripcion_residente = objeto.residente.fullname
            else:
                objeto.descripcion_residente = '-----'

            if objeto.fkinvitado:
                if objeto.fkconductor:
                    objeto.descripcion_nombre_conductor = objeto.conductor.fullname
                else:
                    objeto.descripcion_nombre_conductor = '-----'

                objeto.descripcion_ci_invitado = objeto.invitado.ci
                objeto.descripcion_nombre_invitado = objeto.invitado.fullname


            else:
                objeto.descripcion_ci_invitado = '-----'
                objeto.descripcion_nombre_invitado = '-----'

            if objeto.fkvehiculo:
                objeto.descripcion_placa = objeto.vehiculo.placa
                objeto.descripcion_tipo = objeto.vehiculo.tipo.nombre
                objeto.descripcion_marca = objeto.vehiculo.marca.nombre
                objeto.descripcion_modelo = objeto.vehiculo.modelo.nombre if objeto.vehiculo.fkmodelo else '-----'
                objeto.descripcion_color = objeto.vehiculo.color.nombre

            else:
                objeto.descripcion_placa = '-----'
                objeto.descripcion_tipo = '-----'
                objeto.descripcion_marca = '-----'
                objeto.descripcion_modelo = '-----'
                objeto.descripcion_color = '-----'

            if objeto.fkdomicilio:
                objeto.descripcion_destino = objeto.domicilio.nombre
                objeto.descripcion_condominio = objeto.domicilio.codigocondominio
            elif objeto.fkareasocial:
                objeto.descripcion_destino = objeto.areasocial.nombre
                objeto.descripcion_condominio = objeto.areasocial.codigocondominio
            else:
                objeto.descripcion_destino = '-----'
                objeto.descripcion_condominio = '-----'

            if objeto.fknropase:
                objeto.descripcion_nropase = objeto.nropase.numero + ' ' + objeto.nropase.tipo
            else:
                objeto.descripcion_nropase = '-----'

            self.db.merge(objeto)
            print(str(cont))
            cont = cont +1

        fecha = BitacoraManager(self.db).fecha_actual()

        print("sincronizar inicio commit "+ fecha.strftime('%d/%m/%Y %H:%M:%S'))
        self.db.commit()
        print("sincronizar fin commit "+ fecha.strftime('%d/%m/%Y %H:%M:%S'))

    def reporte_vehicular_visita(self,diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        domicilio = self.db.query(self.entity).join(Domicilio).filter(
            Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
            func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
                self.entity.tipo == "Vehicular").filter(self.entity.estado == True).order_by(self.entity.fechar.desc())

        return domicilio

    # def reporte_movimientos_vehicular(self,diccionario):
    #
    #     diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
    #     diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')
    #
    #     domicilio = self.db.query(self.entity).join(Domicilio).filter(
    #         Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
    #         func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
    #             self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()
    #
    #     areasocial = self.db.query(self.entity).join(Areasocial).filter(
    #         Areasocial.fkcondominio == diccionario['fkcondominio']).filter(
    #         func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
    #             self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()
    #
    #     for area in areasocial:
    #         domicilio.append(area)
    #
    #     print("retorno de movimientos :"+ str(len(domicilio)))
    #     codigo = BitacoraManager(self.db).generar_codigo()
    #
    #     cname = "movimiento_"+codigo+".xlsx"
    #
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.title = 'a'
    #
    #     indice = 1
    #
    #     ws['A' + str(indice)] = 'ID'
    #     ws['B' + str(indice)] = 'Fecha Registro'
    #     ws['C' + str(indice)] = 'Fecha Ingreso'
    #     ws['D' + str(indice)] = 'Fecha Salida'
    #     ws['E' + str(indice)] = 'Tipo documento'
    #     ws['F' + str(indice)] = 'Nº documento'
    #     ws['G' + str(indice)] = 'Expedido documento'
    #     ws['H' + str(indice)] = 'Nombre Invitado'
    #     ws['I' + str(indice)] = 'Apellidop Invitado'
    #     ws['J' + str(indice)] = 'Apellidom Invitado'
    #     ws['K' + str(indice)] = 'Nº documento conductor'
    #     ws['L' + str(indice)] = 'Nombre Conductor'
    #     ws['M' + str(indice)] = 'Apellidop Conductor'
    #     ws['N' + str(indice)] = 'Apellidom Conductor'
    #     ws['O' + str(indice)] = 'Cant. Pasajeros'
    #     ws['P' + str(indice)] = 'Placa'
    #     ws['Q' + str(indice)] = 'Tipo vehiculo'
    #     ws['R' + str(indice)] = 'Marca'
    #     ws['S' + str(indice)] = 'Modelo'
    #     ws['T' + str(indice)] = 'Color'
    #     ws['U' + str(indice)] = 'Destino'
    #     ws['V' + str(indice)] = 'Autorizacion'
    #     ws['W' + str(indice)] = 'Tipo de pase'
    #     ws['X' + str(indice)] = 'Observacion'
    #     ws['Y' + str(indice)] = 'Tipo'
    #
    #     for i in domicilio:
    #         print(str(i.id))
    #
    #         if i.fkinvitado:
    #             print("movimiento exportar: " + str(i.id))
    #
    #             indice = indice + 1
    #             ws['A' + str(indice)] = i.id
    #             ws['B' + str(indice)] = i.fechar
    #             ws['C' + str(indice)] = i.fechai
    #             ws['D' + str(indice)] = i.fechaf
    #             ws['E' + str(indice)] = i.fktipodocumento
    #
    #             ws['F' + str(indice)] = i.invitado.ci
    #             ws['G' + str(indice)] = i.invitado.expendido
    #             ws['H' + str(indice)] = i.invitado.nombre
    #             ws['I' + str(indice)] = i.invitado.apellidop
    #             ws['J' + str(indice)] = i.invitado.apellidom
    #
    #             if i.fkconductor:
    #
    #                 ws['K' + str(indice)] = i.conductor.ci
    #                 ws['L' + str(indice)] = i.conductor.nombre
    #                 ws['M' + str(indice)] = i.conductor.apellidop
    #                 ws['N' + str(indice)] = i.conductor.apellidom
    #
    #             else:
    #                 ws['K' + str(indice)] = None
    #                 ws['L' + str(indice)] = None
    #                 ws['M' + str(indice)] = None
    #                 ws['N' + str(indice)] = None
    #
    #             ws['O' + str(indice)] = i.cantpasajeros
    #             if i.tipo == "Vehicular":
    #
    #                 ws['P' + str(indice)] = i.vehiculo.placa
    #                 ws['Q' + str(indice)] = i.vehiculo.tipo.nombre
    #                 ws['R' + str(indice)] = i.vehiculo.marca.nombre
    #                 ws['S' + str(indice)] = None
    #                 ws['T' + str(indice)] = i.vehiculo.color.nombre
    #             else:
    #                 ws['P' + str(indice)] = None
    #                 ws['Q' + str(indice)] = None
    #                 ws['R' + str(indice)] = None
    #                 ws['S' + str(indice)] = None
    #                 ws['T' + str(indice)] = None
    #
    #             if i.fkdomicilio:
    #                 ws['U' + str(indice)] = i.domicilio.codigo
    #             elif i.fkareasocial:
    #                 ws['U' + str(indice)] = i.areasocial.codigo
    #             else:
    #                 ws['U' + str(indice)] = ""
    #             ws['V' + str(indice)] = i.autorizacion.nombre
    #             ws['W' + str(indice)] = i.tipopase.nombre
    #             ws['X' + str(indice)] = i.observacion
    #             ws['Y' + str(indice)] = i.tipo
    #
    #     wb.save("server/common/resources/downloads/" + cname)
    #     return cname

    def obtener_x_codigo(self, codigo):
        return self.db.query(self.entity).filter(self.entity.codigo == codigo).first()

    def obtener_x_id(self, id):
        return self.db.query(self.entity).filter(self.entity.id == id).first()

    def obtener_destino(self,idMovimiento):
        mov = self.db.query(self.entity).filter(self.entity.id == idMovimiento).first()
        if mov:

            if mov.fkdomicilio:
                return mov.domicilio
            elif mov.fkareasocial:
                return mov.areasocial
            else:
                return None
        else:
            print("obtener destino mov= null")

            return None

    def get_all(self):
        return self.db.query(self.entity)

    def get_all_by_lastname(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.apellidopaterno.asc()).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.tipo == "Vehicular"))

    def listar_movimiento_dia(self,usuario):

        fecha = datetime.now(pytz.timezone('America/La_Paz'))
        fechahoy = str(fecha.day)+"/"+str(fecha.month)+"/"+str(fecha.year)
        fechahoy = datetime.strptime(fechahoy, '%d/%m/%Y')


        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.fechar.cast(Date) == fechahoy).filter(
                self.entity.tipo == "Vehicular").order_by(self.entity.id.desc()).all()
        else:

            domicilio = self.db.query(self.entity).filter(self.entity.estado == True).join(Domicilio).filter(Domicilio.fkcondominio== usuario.fkcondominio).filter(self.entity.fechar.cast(Date) == fechahoy).filter(
                self.entity.tipo == "Vehicular").order_by(self.entity.id.desc()).all()

            areasocial = self.db.query(self.entity).filter(self.entity.estado == True).join(Areasocial).filter(Areasocial.fkcondominio== usuario.fkcondominio).filter(self.entity.fechar.cast(Date) == fechahoy).filter(
                self.entity.tipo == "Vehicular").order_by(self.entity.id.desc()).all()

            for area in areasocial:
                domicilio.append(area)

            return domicilio

    def list_all_reporte(self):
        return self.db.query(self.entity).filter(self.entity.tipo == "Vehicular").limit(1).all()

    def delay(self, diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        domicilio = dict(objects=self.db.query(self.entity).join(Domicilio).filter(
            Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
            func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
                self.entity.tipo == "Vehicular").filter(self.entity.estado == True))

        # print("retorno de movimientos :"+ str(len(domicilio)))
        # return domicilio



        ss = {'reports': domicilio}

        return ss

    # def reporte_movimientos_vehicular(self,diccionario):
    #
    #     diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
    #     diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')
    #
    #     return dict(objects=self.db.query(self.entity).join(Domicilio).filter(
    #         Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
    #         func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
    #             self.entity.tipo == "Vehicular").filter(self.entity.estado == True))

    def reporte_movimientos_vehicular_visita(self, diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        codigo = BitacoraManager(self.db).generar_codigo()

        cname = "movimiento_" + codigo + ".xlsx"

        movimientos = self.db.query(self.entity).filter(
                self.entity.tipo == "Vehicular").filter(func.date(self.entity.fechar)
                                                        .between(diccionario['fechainicio'], diccionario['fechafin'])) \
            .order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'Fecha Registro'
        ws['C' + str(indice)] = 'Fecha Ingreso'
        ws['D' + str(indice)] = 'Fecha Salida'
        ws['E' + str(indice)] = 'Tipo documento'
        ws['F' + str(indice)] = 'Nº documento'
        ws['G' + str(indice)] = 'Expedido documento'
        ws['H' + str(indice)] = 'Nombre Invitado'
        ws['I' + str(indice)] = 'Apellidop Invitado'
        ws['J' + str(indice)] = 'Apellidom Invitado'
        ws['K' + str(indice)] = 'Nº documento conductor'
        ws['L' + str(indice)] = 'Nombre Conductor'
        ws['M' + str(indice)] = 'Apellidop Conductor'
        ws['N' + str(indice)] = 'Apellidom Conductor'
        ws['O' + str(indice)] = 'Cant. Pasajeros'
        ws['P' + str(indice)] = 'Placa'
        ws['Q' + str(indice)] = 'Tipo vehiculo'
        ws['R' + str(indice)] = 'Marca'
        ws['S' + str(indice)] = 'Modelo'
        ws['T' + str(indice)] = 'Color'
        ws['U' + str(indice)] = 'Destino'
        ws['V' + str(indice)] = 'Autorizacion'
        ws['W' + str(indice)] = 'Tipo de pase'
        ws['X' + str(indice)] = 'Observacion'
        ws['Y' + str(indice)] = 'Tipo'

        for i in movimientos:
            print(str(i.id))

            if i.fkinvitado:
                print("movimiento exportar: " + str(i.id))

                indice = indice + 1
                ws['A' + str(indice)] = i.id
                ws['B' + str(indice)] = i.fechar
                ws['C' + str(indice)] = i.fechai
                ws['D' + str(indice)] = i.fechaf
                ws['E' + str(indice)] = i.fktipodocumento

                ws['F' + str(indice)] = i.invitado.ci
                ws['G' + str(indice)] = i.invitado.expendido
                ws['H' + str(indice)] = i.invitado.nombre
                ws['I' + str(indice)] = i.invitado.apellidop
                ws['J' + str(indice)] = i.invitado.apellidom

                if i.fkconductor:

                    ws['K' + str(indice)] = i.conductor.ci
                    ws['L' + str(indice)] = i.conductor.nombre
                    ws['M' + str(indice)] = i.conductor.apellidop
                    ws['N' + str(indice)] = i.conductor.apellidom

                else:
                    ws['K' + str(indice)] = None
                    ws['L' + str(indice)] = None
                    ws['M' + str(indice)] = None
                    ws['N' + str(indice)] = None

                ws['O' + str(indice)] = i.cantpasajeros
                if i.tipo == "Vehicular":

                    ws['P' + str(indice)] = i.vehiculo.placa
                    ws['Q' + str(indice)] = i.vehiculo.tipo.nombre
                    ws['R' + str(indice)] = i.vehiculo.marca.nombre
                    ws['S' + str(indice)] = None
                    ws['T' + str(indice)] = i.vehiculo.color.nombre
                else:
                    ws['P' + str(indice)] = None
                    ws['Q' + str(indice)] = None
                    ws['R' + str(indice)] = None
                    ws['S' + str(indice)] = None
                    ws['T' + str(indice)] = None

                if i.fkdomicilio:
                    ws['U' + str(indice)] = i.domicilio.codigo
                elif i.fkareasocial:
                    ws['U' + str(indice)] = i.areasocial.codigo
                else:
                    ws['U' + str(indice)] = ""

                ws['V' + str(indice)] = i.autorizacion.nombre
                ws['W' + str(indice)] = i.tipopase.nombre
                ws['X' + str(indice)] = i.observacion
                ws['Y' + str(indice)] = i.tipo

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def reporte_movimientos_peatonal_visita(self, diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        codigo = BitacoraManager(self.db).generar_codigo()

        cname = "movimiento_" + codigo + ".xlsx"

        movimientos = self.db.query(self.entity).filter(
            self.entity.tipo == "Peatonal").filter(func.date(self.entity.fechar)
                                                    .between(diccionario['fechainicio'], diccionario['fechafin'])) \
            .order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'Fecha Registro'
        ws['C' + str(indice)] = 'Fecha Ingreso'
        ws['D' + str(indice)] = 'Fecha Salida'
        ws['E' + str(indice)] = 'Tipo documento'
        ws['F' + str(indice)] = 'Nº documento'
        ws['G' + str(indice)] = 'Expedido documento'
        ws['H' + str(indice)] = 'Nombre Invitado'
        ws['I' + str(indice)] = 'Apellidop Invitado'
        ws['J' + str(indice)] = 'Apellidom Invitado'
        ws['K' + str(indice)] = 'Destino'
        ws['L' + str(indice)] = 'Autorizacion'
        ws['M' + str(indice)] = 'Tipo de pase'
        ws['N' + str(indice)] = 'Observacion'
        ws['O' + str(indice)] = 'Tipo'

        for i in movimientos:
            print(str(i.id))

            if i.fkinvitado:
                print("movimiento exportar: " + str(i.id))

                indice = indice + 1
                ws['A' + str(indice)] = i.id
                ws['B' + str(indice)] = i.fechar
                ws['C' + str(indice)] = i.fechai
                ws['D' + str(indice)] = i.fechaf
                ws['E' + str(indice)] = i.fktipodocumento

                ws['F' + str(indice)] = i.invitado.ci
                ws['G' + str(indice)] = i.invitado.expendido
                ws['H' + str(indice)] = i.invitado.nombre
                ws['I' + str(indice)] = i.invitado.apellidop
                ws['J' + str(indice)] = i.invitado.apellidom


                if i.fkdomicilio:
                    ws['K' + str(indice)] = i.domicilio.codigo
                elif i.fkareasocial:
                    ws['K' + str(indice)] = i.areasocial.codigo
                else:
                    ws['K' + str(indice)] = ""

                ws['L' + str(indice)] = i.autorizacion.nombre
                ws['M' + str(indice)] = i.tipopase.nombre
                ws['N' + str(indice)] = i.observacion
                ws['O' + str(indice)] = i.tipo

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def reporte_movimientos_vehicular(self,diccionario):

        diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
        diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')

        domicilio = self.db.query(self.entity).join(Domicilio).filter(
            Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
            func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
                self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()

        areasocial = self.db.query(self.entity).join(Areasocial).filter(
            Areasocial.fkcondominio == diccionario['fkcondominio']).filter(
            func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
                self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()



        for area in areasocial:
            domicilio.append(area)

        print("retorno de movimientos :"+ str(len(domicilio)))
        return domicilio

    # def reporte_movimientos_vehicular(self,diccionario):
    #
    #     diccionario['fechainicio'] = datetime.strptime(diccionario['fechainicio'], '%d/%m/%Y')
    #     diccionario['fechafin'] = datetime.strptime(diccionario['fechafin'], '%d/%m/%Y')
    #
    #
    #
    #     domicilio = self.db.query(self.entity).join(Domicilio).filter(
    #         Domicilio.fkcondominio == diccionario['fkcondominio']).filter(
    #         func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
    #             self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()
    #
    #     areasocial = self.db.query(self.entity).join(Areasocial).filter(
    #         Areasocial.fkcondominio == diccionario['fkcondominio']).filter(
    #         func.date(self.entity.fechar).between(diccionario['fechainicio'], diccionario['fechafin'])).filter(
    #             self.entity.tipo == "Vehicular").filter(self.entity.estado == True).all()
    #
    #
    #     for area in areasocial:
    #         domicilio.append(area)
    #
    #
    #
    #
    #     print("retorno de movimientos :"+ str(len(domicilio)))
    #
    #     def _hilo(domicilio):
    #         cont = 1
    #         print("hiol")
    #         list = []
    #         for mov in domicilio:
    #             print(str(cont))
    #             print(str(mov.id))
    #             print(str(mov.vehiculo.tipo))
    #
    #             list.append(dict(id=mov.id, fechai=mov.fechai.strftime('%d/%m/%Y') if mov.fechai else '----',
    #                              fechaf=mov.fechaf.strftime('%d/%m/%Y') if mov.fechaf else '----',
    #                              tipodocumento=mov.tipodocumento.nombre, ci_invitado=mov.invitado.ci,
    #                              nombre_invitado=mov.invitado.fullname,
    #                              nombre_conductor=mov.conductor.fullname if mov.fkconductor else '',
    #                              cantpasajeros=mov.cantpasajeros, placa=mov.vehiculo.placa, tipo_vehiculo='',
    #                              marca=mov.vehiculo.marca.nombre,
    #                              modelo=mov.vehiculo.modelo.nombre if mov.vehiculo.fkmodelo else '',
    #                              color=mov.vehiculo.color.nombre,
    #                              destino=mov.domicilio.nombre if mov.fkdomicilio else 'Area Social',
    #                              autorizacion=mov.autorizacion.nombre, nropase=mov.nropase.tipo,
    #                              tipopase=mov.tipopase.nombre, observacion=mov.observacion))
    #             cont = cont + 1
    #         print("registros: " + str(len(domicilio)))
    #
    #         return list
    #
    #     t1 = threading.Thread(name="hilo",target=_hilo, args=(domicilio, ))
    #
    #     t1.start()
    #
    #     list = t1.join()
    #
    #     return list

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()

    def insert(self, diccionary):

        diccionary['cantpasajeros'] = abs(int(diccionary['cantpasajeros']))

        diccionary['placa'] = diccionary['placa'].replace(" ", "")

        if diccionary['fkinvitacion'] == "":
            diccionary['fkinvitacion'] = None

        accesos_invitacion = InvitacionManager(self.db).obtener_accesos_evento(diccionary['fkinvitacion'])

        if accesos_invitacion['paselibre']:
            diccionary['fkvehiculo'] = None
            diccionary['fkconductor'] = None
            diccionary['fkinvitado'] = None
        else:

            if diccionary['visita']:
                if diccionary['fkinvitado'] == "" or diccionary['fkinvitado'] == "0":
                    if diccionary['ci'] != "":
                        invitado = InvitadoManager(self.db).registrar_invitado_movimiento(diccionary)
                        diccionary['fkinvitado'] = invitado.id
                    else:
                        diccionary['fkinvitado'] = None

                else:
                    invitado = InvitadoManager(self.db).actualizar_invitado(diccionary)
            else:
                diccionary['fkinvitado'] = None


            if diccionary['fkconductor'] == "" or diccionary['fkconductor'] == "0":
                diccionary['nombre_conductor'] = diccionary['nombre_conductor'].replace(" ", "")
                if diccionary['nombre_conductor'] != "":
                    conductor = InvitadoManager(self.db).registrar_conductor(diccionary)
                    diccionary['fkconductor'] = conductor.id
                else:
                    diccionary['fkconductor'] = None
            else:
                conductor = InvitadoManager(self.db).registrar_conductor(diccionary)


            if diccionary['fkvehiculo'] == "" or diccionary['fkvehiculo'] == "0":
                vehiculo = VehiculoManager(self.db).registrar_vehiculo(diccionary)
                diccionary['fkvehiculo'] = vehiculo


        if diccionary['fktipodocumento_conductor'] == "":
            diccionary['fktipodocumento_conductor'] = None

        if diccionary['fkdomicilio'] == "":
            diccionary['fkdomicilio'] = None

        if diccionary['fkareasocial'] == "":
            diccionary['fkareasocial'] = None

        if diccionary['fkmodelo'] == "":
            diccionary['fkmodelo'] = None

        if diccionary['fkmarca'] == "":
            diccionary['fkmarca'] = None
        elif diccionary['fkmarca'] == "0":
            diccionary['fkmarca'] = None

        if diccionary['cantpasajeros'] == "":
            diccionary['cantpasajeros'] = None

        try:
            if diccionary['fkresidente'] == "":
                diccionary['fkresidente'] = None

        except Exception as e:
            print("no se envio fkresidente")
            diccionary['fkresidente'] = None

        fecha = BitacoraManager(self.db).fecha_actual()

        diccionary['tipo'] = "Vehicular"
        diccionary['descripcion_ci_invitado'] = diccionary['ci']+" "+ diccionary['apellidop']
        diccionary['descripcion_nombre_invitado'] = diccionary['nombre'] +" "+ diccionary['apellidop'] +" "+diccionary['apellidom']

        # diccionary['fechai'] = fecha
        if diccionary['fechar'] == "":
            diccionary['fechar'] = fecha

        print("fechar: " + str(diccionary['fechar']))
        objeto = MovimientoManager(self.db).entity(**diccionary)

        nropase = self.db.query(Nropase).filter(Nropase.id == objeto.fknropase).filter(Nropase.situacion == "Ocupado").first()

        if nropase :
            print("bloqueo de Registro duplicado: tarjeta ocupada " + nropase.tipo+" "+nropase.numero )
            b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Bloqueo de registro duplicado.", fecha=fecha,tabla="movimiento", identificador=0)
            super().insert(b)
            return False
        else:
            a = super().insert(objeto)

            print("registro ingreso Vehicular: " +str(a.id))
            b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Movimiento.", fecha=fecha,tabla="movimiento", identificador=a.id)
            super().insert(b)


            if a.fknropase:
                # actualizar siuacion
                NropaseManager(self.db).situacion(a.fknropase, "Ocupado")

            # deshabilitar invitacion
            if a.fkinvitacion:

                if accesos_invitacion['multiacceso'] is False:
                    if accesos_invitacion['multiple'] is False:
                        #if accesos_invitacion['paselibre'] is False:
                         InvitacionManager(self.db).delete(a.fkinvitacion, False, objeto.user, objeto.ip)

                principal = self.db.query(Principal).first()
                if principal.estado:
                    NotificacionManager(self.db).registrar_notificacion_(a,objeto)

            self.hilo_sincronizar_update_descripcion(a)

            # t = Thread(target=self.hilo_sincronizar_update_descripcion, args=(a,))
            # t.start()


            return a

    def hilo_sincronizar_update_descripcion(self, objeto):
        if objeto.codigo == "":
            objeto.codigo = objeto.id

        objeto.descripcion_documento = objeto.tipodocumento.nombre

        if objeto.fkresidente:
            objeto.descripcion_residente = objeto.residente.fullname
        else:
            objeto.descripcion_residente = '-----'


        if objeto.fkinvitado:
            if objeto.fkconductor:
                objeto.descripcion_nombre_conductor = objeto.conductor.fullname
            else:
                objeto.descripcion_nombre_conductor = '-----'

            objeto.descripcion_ci_invitado = objeto.invitado.ci
            objeto.descripcion_nombre_invitado = objeto.invitado.fullname

        else:
            objeto.descripcion_ci_invitado = '-----'
            objeto.descripcion_nombre_invitado =  '-----'

        if objeto.fkvehiculo:
            objeto.descripcion_placa = objeto.vehiculo.placa
            objeto.descripcion_tipo = objeto.vehiculo.tipo.nombre
            objeto.descripcion_marca = objeto.vehiculo.marca.nombre
            objeto.descripcion_modelo = objeto.vehiculo.modelo.nombre if objeto.vehiculo.fkmodelo else '-----'
            objeto.descripcion_color = objeto.vehiculo.color.nombre


        else:
            objeto.descripcion_placa = '-----'
            objeto.descripcion_tipo = '-----'
            objeto.descripcion_marca = '-----'
            objeto.descripcion_modelo = '-----'
            objeto.descripcion_color = '-----'

        if objeto.fkdomicilio:
            objeto.descripcion_destino = objeto.domicilio.nombre
            objeto.descripcion_condominio = objeto.domicilio.codigocondominio
        elif objeto.fkareasocial:
            objeto.descripcion_destino = objeto.areasocial.nombre
            objeto.descripcion_condominio = objeto.areasocial.codigocondominio
        else:
            objeto.descripcion_destino = '-----'

        if objeto.fknropase:
            objeto.descripcion_nropase = objeto.nropase.numero + ' ' + objeto.nropase.tipo

        self.db.merge(objeto)
        self.db.commit()

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()
        objeto.fechanacimiento = datetime.strptime(objeto.fechanacimiento, '%d/%m/%Y')

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Movimiento.", fecha=fecha,tabla="residente", identificador=a.id)
        super().insert(b)
        return a

    def salida(self, id, user, ip):
        x = self.db.query(Movimiento).filter(Movimiento.id == id).first()
        # fecha = BitacoraManager(self.db).fecha_actual()
        fecha = datetime.now(pytz.timezone('America/La_Paz'))

        print("fecha salida: "+ str(fecha))
        if x.fechai is None:
            x.fechai = x.fechar
            x.descripcion_fechai = x.fechar.strftime('%d/%m/%Y %H:%M:%S')


        x.fechaf = fecha
        x.descripcion_fechaf = fecha.strftime('%d/%m/%Y %H:%M:%S')

        b = Bitacora(fkusuario=user, ip=ip, accion="Registro Salida", fecha=fecha, tabla="movimiento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()


        if x.fknropase:
            # actualizar siuacion
            NropaseManager(self.db).situacion(x.fknropase, "Libre")
        return x

    def asignar_codigo(self, id, codigo):
        x = self.db.query(Movimiento).filter(Movimiento.id == id).first()

        if x:

            x.codigo = codigo

            self.db.merge(x)
            self.db.commit()
            print("codigo asignado: " + str(x.codigo))


        return x

    def delete(self, id, user, ip):
        x = self.db.query(Movimiento).filter(Movimiento.id == id).first()
        x.estado = False

        mensaje = "Deshabilito Movimiento"

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion=mensaje, fecha=fecha, tabla="movimiento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        principal = self.db.query(Principal).first()

        if principal.estado:
            condominio = self.db.query(Condominio).filter(Condominio.codigo == x.descripcion_condominio).first()
            try:
                if condominio:

                    if condominio.ip_publica != "":
                        diccionary = dict(id=id, estado=False, user=user, ip=ip)

                        url = "http://" + condominio.ip_publica + ":" + condominio.puerto + "/api/v1/sincronizar_movimiento_estado"

                        headers = {'Content-Type': 'application/json'}
                        string = diccionary
                        cadena = json.dumps(string)
                        body = cadena
                        resp = requests.post(url, data=body, headers=headers, verify=False)
                        response = json.loads(resp.text)

                        print(response)


            except Exception as e:
                # Other errors are possible, such as IOError.
                print("Error de conexion: " + str(e))

        return x

    def salida_sincronizada_nube(self, id, fechaf, user, ip):
        print("salida nube")
        x = self.db.query(Movimiento).filter(Movimiento.id == id).first()
        fecha = BitacoraManager(self.db).fecha_actual()
        if x.fechai is None:
            x.fechai = x.fechar
            x.descripcion_fechai = x.fechar.strftime('%d/%m/%Y %H:%M:%S')

        x.fechaf = datetime.strptime(fechaf, '%d/%m/%Y %H:%M:%S')
        x.descripcion_fechaf = fechaf


        b = Bitacora(fkusuario=user, ip=ip, accion="Registro Salida sincronizada nube", fecha=fecha, tabla="movimiento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        if x.fknropase:
            # actualizar siuacion
            NropaseManager(self.db).situacion(x.fknropase, "Libre")

        return x

    def salida_sincronizada(self, id, fechaf, user, ip):
        print("salida")
        x = self.db.query(Movimiento).filter(Movimiento.id == id).first()
        fecha = BitacoraManager(self.db).fecha_actual()
        if x.fechai is None:
            x.fechai = x.fechar
            x.descripcion_fechai = x.fechar.strftime('%d/%m/%Y %H:%M:%S')

        x.fechaf = fechaf
        x.descripcion_fechaf = fechaf.strftime('%d/%m/%Y %H:%M:%S')


        b = Bitacora(fkusuario=user, ip=ip, accion="Registro Salida", fecha=fecha, tabla="movimiento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        if x.fknropase:
            # actualizar siuacion
            NropaseManager(self.db).situacion(x.fknropase, "Libre")

        return x


    def armado_diccionario(self, d):
        return dict(id=d.id, fechar=d.fechar.strftime('%d/%m/%Y %H:%M:%S'),
                          fechai=d.descripcion_fechai, fechaf=d.descripcion_fechaf,
                          documento=d.descripcion_documento,
                          ci_invitado=d.descripcion_ci_invitado,
                          nombre_invitado=d.descripcion_nombre_invitado,
                          nombre_conductor=d.descripcion_nombre_conductor,
                          cantpasajeros=d.cantpasajeros,
                          placa=d.descripcion_placa, tipo=d.descripcion_tipo,
                          marca=d.descripcion_marca,
                          residente=d.descripcion_residente,
                          color=d.descripcion_color, destino=d.descripcion_destino,
                          autorizacion=d.autorizacion.nombre,
                          nropase=d.descripcion_nropase, tipopase=d.tipopase.nombre,
                          observacion=d.observacion)


    def filtrar(self, fechainicio, fechafin,usuario):
        usuario = UsuarioManager(self.db).get_by_pass(usuario)
        lista = list()

        print("FIltrar por fecha")

        if usuario.sigas:
            movimientos = self.db.query(self.entity)\
                .filter(self.entity.estado == True)\
                .filter(func.date(self.entity.fechar)
                .between(fechainicio, fechafin))\
                .filter(self.entity.tipo == "Vehicular").all()

            for d in movimientos:
                lista.append(self.armado_diccionario(d))

            print("retorno de movimientos :" + str(len(lista)))
            return lista

        else:
            movimientos = self.db.query(self.entity)\
                .filter(self.entity.descripcion_condominio== usuario.condominio.codigo)\
                .filter(func.date(self.entity.fechar).between(fechainicio, fechafin))\
                .filter(self.entity.tipo == "Vehicular")\
                .filter(self.entity.estado == True).all()


            for d in movimientos:
                lista.append(self.armado_diccionario(d))

            print("retorno de movimientos :" + str(len(lista)))
            return lista

    def filtrar_residente(self, fechainicio, fechafin, fresidente):
        lista = list()
        c = 0
        fecha = fecha_zona
        fechahoy = str(fecha.day) + "/" + str(fecha.month) + "/" + str(fecha.year)
        fechahoy = datetime.strptime(fechahoy, '%d/%m/%Y')

        movimientos = self.db.query(self.entity).filter(self.entity.estado == True) \
            .filter(self.entity.fkresidente == fresidente) \
            .filter(func.date(self.entity.fechar).between(fechainicio, fechafin)) \
            .filter(self.entity.tipo == "Vehicular").all()

        for d in movimientos:
            lista.append(self.armado_diccionario(d))

        print("retorno de movimientos :" + str(len(lista)))
        return lista

    def filtrar_domicilio(self, fechainicio, fechafin, fdomicilio):
        lista = list()
        c = 0
        fecha = fecha_zona
        fechahoy = str(fecha.day) + "/" + str(fecha.month) + "/" + str(fecha.year)
        fechahoy = datetime.strptime(fechahoy, '%d/%m/%Y')

        movimientos = self.db.query(self.entity) \
            .filter(self.entity.estado == True) \
            .filter(self.entity.fkdomicilio == fdomicilio) \
            .filter(self.entity.tipo == "Vehicular") \
            .filter(func.date(self.entity.fechar).between(fechainicio, fechafin)).all()

        for d in movimientos:
            lista.append(self.armado_diccionario(d))

        print("retorno de movimientos :" + str(len(lista)))
        return lista



    def filtrar_movil(self, fechainicio, fechafin,usuario):

        list = {}
        c = 0


        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(func.date(self.entity.fechar).between(fechainicio, fechafin)).all()
        else:
            domicilio = self.db.query(self.entity).filter(self.entity.estado == True).join(Domicilio).filter(Domicilio.fkcondominio== usuario.fkcondominio).filter(func.date(self.entity.fechar).between(fechainicio, fechafin)).all()

            areasocial = self.db.query(self.entity).filter(self.entity.estado == True).join(Areasocial).filter(Areasocial.fkcondominio== usuario.fkcondominio).filter(func.date(self.entity.fechar).between(fechainicio, fechafin)).all()

            for area in areasocial:
                domicilio.append(area)

            return domicilio

    def actualizar_movimiento(self, marcacion):

        nropase = self.db.query(Nropase).filter(Nropase.tarjeta == marcacion.tarjeta).first()

        if nropase:

            if nropase.tipo == "Excepcion":
                mov = self.db.query(Movimiento).join(Nropase).filter(Movimiento.estado == True).filter(Nropase.tarjeta == marcacion.tarjeta)\
                    .filter(Movimiento.fechai == None).first()

            else:

                mov = self.db.query(Movimiento).join(Nropase).filter(Movimiento.estado == True).filter(Nropase.tarjeta == marcacion.tarjeta).filter(
                        or_(Movimiento.fechai == None,Movimiento.fechaf == None)).first()

        else:
            mov = self.db.query(Movimiento).join(Nropase).filter(Movimiento.estado == True).filter(
                Nropase.tarjeta == marcacion.tarjeta).filter(
                or_(Movimiento.fechai == None, Movimiento.fechaf == None)).first()

        if mov:
            if not mov.fechai:
                mov.fechai = marcacion.time
                mov.descripcion_fechai = marcacion.time.strftime('%d/%m/%Y %H:%M:%S')

                # if mov.nropase.tipo == "Excepcion":
                # mov.fechaf = marcacion.time

                self.db.merge(mov)
                self.db.commit()
            elif not mov.fechaf:
                mov.fechaf = marcacion.time
                mov.descripcion_fechaf = marcacion.time.strftime('%d/%m/%Y %H:%M:%S')
                self.db.merge(mov)

                nropase = self.db.query(Nropase).filter(Nropase.id == mov.fknropase).first()
                nropase.situacion = "Libre"
                self.db.merge(nropase)
                self.db.commit()

        marcacion.sincronizado =True

        return marcacion

    def movimiento_excel(self, fechainicio, fechafin):
        fechainicio = datetime.strptime(fechainicio, '%d/%m/%Y')
        fechafin = datetime.strptime(fechafin, '%d/%m/%Y')

        codigo = BitacoraManager(self.db).generar_codigo()

        cname = "movimiento_"+codigo+".xlsx"

        movimientos = self.db.query(self.entity).filter(func.date(self.entity.fechar)
                                                .between(fechainicio, fechafin))\
                                                .order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'Fecha Registro'
        ws['C' + str(indice)] = 'Fecha Ingreso'
        ws['D' + str(indice)] = 'Fecha Salida'
        ws['E' + str(indice)] = 'Tipo documento'
        ws['F' + str(indice)] = 'Nº documento'
        ws['G' + str(indice)] = 'Expedido documento'
        ws['H' + str(indice)] = 'Nombre Invitado'
        ws['I' + str(indice)] = 'Apellidop Invitado'
        ws['J' + str(indice)] = 'Apellidom Invitado'
        ws['K' + str(indice)] = 'Nº documento conductor'
        ws['L' + str(indice)] = 'Nombre Conductor'
        ws['M' + str(indice)] = 'Apellidop Conductor'
        ws['N' + str(indice)] = 'Apellidom Conductor'
        ws['O' + str(indice)] = 'Cant. Pasajeros'
        ws['P' + str(indice)] = 'Placa'
        ws['Q' + str(indice)] = 'Tipo vehiculo'
        ws['R' + str(indice)] = 'Marca'
        ws['S' + str(indice)] = 'Modelo'
        ws['T' + str(indice)] = 'Color'
        ws['U' + str(indice)] = 'Destino'
        ws['V' + str(indice)] = 'Autorizacion'
        ws['W' + str(indice)] = 'Tipo de pase'
        ws['X' + str(indice)] = 'Observacion'
        ws['Y' + str(indice)] = 'Tipo'

        for i in movimientos:
            print(str(i.id))

            if i.fkinvitado:
                print("movimiento exportar: "+ str(i.id))

                indice = indice + 1
                ws['A' + str(indice)] = i.id
                ws['B' + str(indice)] = i.fechar
                ws['C' + str(indice)] = i.fechai
                ws['D' + str(indice)] = i.fechaf
                ws['E' + str(indice)] = i.fktipodocumento

                ws['F' + str(indice)] = i.invitado.ci
                ws['G' + str(indice)] = i.invitado.expendido
                ws['H' + str(indice)] = i.invitado.nombre
                ws['I' + str(indice)] = i.invitado.apellidop
                ws['J' + str(indice)] = i.invitado.apellidom

                if i.fkconductor:

                    ws['K' + str(indice)] = i.conductor.ci
                    ws['L' + str(indice)] = i.conductor.nombre
                    ws['M' + str(indice)] = i.conductor.apellidop
                    ws['N' + str(indice)] = i.conductor.apellidom

                else:
                    ws['K' + str(indice)] = None
                    ws['L' + str(indice)] = None
                    ws['M' + str(indice)] = None
                    ws['N' + str(indice)] = None

                ws['O' + str(indice)] = i.cantpasajeros
                if i.tipo == "Vehicular":

                    ws['P' + str(indice)] = i.vehiculo.placa
                    ws['Q' + str(indice)] = i.vehiculo.tipo.nombre
                    ws['R' + str(indice)] = i.vehiculo.marca.nombre
                    ws['S' + str(indice)] = None
                    ws['T' + str(indice)] = i.vehiculo.color.nombre
                else:
                    ws['P' + str(indice)] = None
                    ws['Q' + str(indice)] = None
                    ws['R' + str(indice)] = None
                    ws['S' + str(indice)] = None
                    ws['T' + str(indice)] = None


                if i.fkdomicilio:
                    ws['U' + str(indice)] = i.domicilio.codigo
                elif i.fkareasocial:
                    ws['U' + str(indice)] = i.areasocial.codigo
                else:
                    ws['U' + str(indice)] = ""
                ws['V' + str(indice)] = i.autorizacion.nombre
                ws['W' + str(indice)] = i.tipopase.nombre
                ws['X' + str(indice)] = i.observacion
                ws['Y' + str(indice)] = i.tipo


        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'Fecha Registro', 'Fecha Ingreso', 'Fecha Salida', 'Tipo documento', 'Nº documento', 'Expedido documento', 'Nombre Invitado', 'Apellidop Invitado',
                        'Apellidom Invitado',
                        'Nº documento conductor',
                        'Nombre Conductor',
                        'Apellidop Conductor',
                        'Apellidom Conductor',
                        'Cant. Pasajeros',
                        'Placa',
                        'Tipo vehiculo',
                        'Marca',
                        'Modelo',
                        'Color',
                        'Destino',
                        'Autorizacion',
                        'Tipo de pase',
                        'Observacion',
                        'Tipo']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):

                    id = row[indices['ID']].value

                    ci = row[indices['Nº documento']].value
                    expedido = row[indices['Expedido documento']].value
                    nombre = row[indices['Nombre Invitado']].value
                    apellidop = row[indices['Apellidop Invitado']].value
                    apellidom = row[indices['Apellidom Invitado']].value

                    ci_conductor = row[indices['Nº documento conductor']].value
                    nombre_conductor = row[indices['Nombre Conductor']].value
                    apellidop_conductor = row[indices['Apellidop Conductor']].value
                    apellidom_conductor = row[indices['Apellidom Conductor']].value


                    if row[indices['ID']].value is not None:

                        print("movimiento import: " + str(row[indices['ID']].value))

                        query_domicilio = self.db.query(Domicilio).filter(Domicilio.codigo == row[indices['Destino']].value).first()

                        if query_domicilio:
                            result_domicilio =query_domicilio.id
                            result_areasocial = None
                        else:
                            query_areasocial = self.db.query(Areasocial).filter(Areasocial.codigo == row[indices['Destino']].value).first()

                            if query_areasocial:
                                result_domicilio = None
                                result_areasocial = query_areasocial.id

                            else:
                                result_domicilio = None
                                result_areasocial = None


                        # query = self.db.query(self.entity).filter(self.entity.fechar == str(row[indices['Fecha Registro']].value)) \
                        #     .filter(self.entity.fkdomicilio == str(result_domicilio)).filter(self.entity.fkareasocial == str(result_areasocial)).all()


                        fkinvitado = None
                        fkconductor = None
                        fkvehiculo = None


                        query_invitado = self.db.query(Invitado).filter(and_(Invitado.ci == ci,Invitado.nombre == nombre,Invitado.apellidop== apellidop,Invitado.apellidom == apellidom)).first()

                        if query_invitado :
                            print("existe invitado")
                            fkinvitado = query_invitado.id

                        else:
                            invi = Invitado(nombre=nombre, apellidop=apellidop, apellidom=apellidom,
                                            ci=ci, expendido=expedido)

                            self.db.add(invi)
                            self.db.flush()

                            fkinvitado = invi.id


                        if nombre_conductor:
                            query_conductor = self.db.query(Invitado).filter(and_(Invitado.ci == ci_conductor,Invitado.nombre == nombre_conductor,Invitado.apellidop== apellidop_conductor,
                                                                                  Invitado.apellidom == apellidom_conductor)).first()

                            if query_conductor :
                                print("existe conductor")
                                fkconductor = query_conductor.id

                            else:
                                condu = Invitado(nombre=nombre_conductor, apellidop=apellidop_conductor, apellidom=apellidom_conductor,
                                                ci=ci_conductor, expendido=expedido)

                                self.db.add(condu)
                                self.db.flush()

                                fkconductor = condu.id

                        if row[indices['Placa']].value:

                            dict_vehiculo = VehiculoManager(self.db).obtener_vehiculo(row[indices['Placa']].value, row[indices['Color']].value, row[indices['Tipo vehiculo']].value, row[indices['Marca']].value,
                                                                                      row[indices['Modelo']].value, '')

                            if dict_vehiculo['id'] != "":
                                print("existe vehiculo")
                                fkvehiculo = dict_vehiculo['id']

                            else:
                                dict_vehiculo['id'] = None

                                vehi = Vehiculo(placa=dict_vehiculo['placa'], fkcolor=dict_vehiculo['fkcolor'], fktipo=dict_vehiculo['fktipo'], fkmarca=dict_vehiculo['fkmarca'], fkmodelo=dict_vehiculo['fkmodelo'])

                                self.db.add(vehi)
                                self.db.flush()

                                fkvehiculo = vehi.id





                        query_domicilio = self.db.query(Domicilio).filter(Domicilio.codigo == row[indices['Destino']].value).first()

                        query_tipo_pase = self.db.query(Tipopase).filter(Tipopase.nombre == row[indices['Tipo de pase']].value).first()

                        query_autorizacion = self.db.query(Autorizacion).filter(Autorizacion.nombre == row[indices['Autorizacion']].value).first()

                        movi = Movimiento(fktipodocumento=row[indices['Tipo documento']].value,
                                          fkinvitado=fkinvitado, fkvehiculo=fkvehiculo,
                                          fechai=row[indices['Fecha Ingreso']].value, fechaf=row[indices['Fecha Salida']].value, fechar=row[indices['Fecha Registro']].value, fkautorizacion=query_autorizacion.id,
                                          fkdomicilio=result_domicilio, fkareasocial=result_areasocial, fktipopase=query_tipo_pase.id, observacion=row[indices['Observacion']].value, tipo=row[indices['Tipo']].value,
                                          cantpasajeros=row[indices['Cant. Pasajeros']].value,fktipodocumento_conductor=None,fkconductor=fkconductor)

                        self.db.merge(movi)
                        self.db.flush()

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def recargar(self, fechainicio, fechafin,usuario,ult_registro):
        condominio = self.db.query(Condominio).filter(Condominio.id == usuario.fkcondominio).first()
        list = {}
        c = 0

        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id > ult_registro).filter(func.date(self.entity.fechar).between(fechainicio, fechafin)).filter(
                self.entity.tipo == "Vehicular").order_by(self.entity.id.desc()).all()
        else:
            domicilio = self.db.query(self.entity)\
                .filter(self.entity.id > ult_registro).filter(self.entity.descripcion_condominio== condominio.codigo)\
                .filter(func.date(self.entity.fechar).between(fechainicio, fechafin))\
                .filter(self.entity.tipo == "Vehicular")\
                .filter(self.entity.estado == True)\
                .order_by(self.entity.id.desc()).all()

            return domicilio


class TipopaseManager(SuperManager):
    def __init__(self, db):
        super().__init__(Tipopase, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()


class TipodocumentoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Tipodocumento, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()


class AutorizacionManager(SuperManager):
    def __init__(self, db):
        super().__init__(Autorizacion, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.id.asc()).all()