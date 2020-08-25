from .models import *
from ...common.managers import *
from sqlalchemy.exc import IntegrityError
from ...common.managers import *
from ...operaciones.bitacora.managers import *
from ...usuarios.usuario.managers import *
from ..residente.managers import *
from ..modelo.models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font



class MarcaManager(SuperManager):

    def __init__(self, db):
        super().__init__(Marca, db)


    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.estado == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.nombre.asc()).all()


    def insert(self, diccionary):

        objeto = MarcaManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Marca.", fecha=fecha,tabla="marca", identificador=a.id)
        super().insert(b)
        return a

    def update(self, diccionary):
        objeto = MarcaManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Marca.", fecha=fecha,tabla="marca", identificador=a.id)
        super().insert(b)
        return a

    def delete(self, id,estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.estado = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Marca.", fecha=fecha, tabla="marca", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        return x

    def obtener_x_nombre(self,marca):
        x =  self.db.query(self.entity).filter(self.entity.nombre == marca).first()

        if x:
            return x.id
        else:
            marc = Marca(nombre=marca)
            m = super().insert(marc)
            return m.id


    def importar_excel(self, cname,user,ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['MARCA','MODELO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    marca = row[indices['MARCA']].value
                    modelo = row[indices['MODELO']].value

                    if marca is not None:

                        query = self.db.query(Modelo).filter(Modelo.nombre == str(modelo)).first()

                        if not query:
                            idmarca = MarcaManager(self.db).obtener_x_nombre(marca)

                            mode = Modelo(nombre=str(modelo),fkmarca=idmarca)

                            self.db.merge(mode)
                            self.db.flush()
                        else:
                            idmarca = MarcaManager(self.db).obtener_x_nombre(marca)

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}

