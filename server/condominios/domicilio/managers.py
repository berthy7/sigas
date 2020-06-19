from ...operaciones.bitacora.managers import *
from ...usuarios.usuario.managers import *
from server.common.managers import SuperManager
from .models import *
from sqlalchemy import or_
from sqlalchemy import and_
from sqlalchemy.exc import IntegrityError
from ..condominio.models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


class DomicilioManager(SuperManager):

    def __init__(self, db):
        super().__init__(Domicilio, db)

    def obtener_x_nombre(self, nombre):
        return self.db.query(self.entity).filter(self.entity.nombre == nombre).first()

    def obtener_x_id(self,id):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == id).first()

    def obtener_fkcondominio(self,iddomicilio):
        x=  self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == iddomicilio).first()
        return x.fkcondominio

    def listar_domicilios(self,usuario):

        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.nombre.asc()).all()

        if usuario.rol.nombre != "RESIDENTE":
            return self.db.query(self.entity).filter(self.entity.fkcondominio == usuario.fkcondominio).filter(self.entity.estado == True).order_by(self.entity.nombre.asc()).all()

        else:
            return None

    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.estado == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.tipo.desc(),self.entity.ubicacion.asc()).all()

    def listar_casas(self,usuario):

        if usuario.fkcondominio:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.fkcondominio == usuario.fkcondominio).filter(self.entity.tipo == "Casa").order_by(self.entity.id.asc()).all()
        elif usuario.fkresidente:
            return self.db.query(self.entity).join(ResidenteDomicilio).join(Residente).filter(Residente.id == usuario.fkresidente).filter(self.entity.estado == True).order_by(self.entity.id.asc()).all()

        else:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.tipo == "Casa").order_by(self.entity.id.asc()).all()

    def listar_departamentos(self,usuario):

        if usuario.fkcondominio:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(
                self.entity.fkcondominio == usuario.fkcondominio).filter(self.entity.tipo == "Departamento").all()
        elif usuario.fkresidente:
            x = self.db.query(Domicilio).join(ResidenteDomicilio).filter(ResidenteDomicilio.vivienda == True).filter(ResidenteDomicilio.fkresidente == usuario.fkresidente).first()

            return self.db.query(Domicilio).filter(Domicilio.tipo == "Departamento").filter(Domicilio.fkcondominio== x.condominio.id).all()

        else:
            return self.db.query(self.entity).filter(self.entity.estado == True).filter(
                self.entity.tipo == "Departamento").all()

    def listar_x_sucursal(self, idsurcusal):
        return self.db.query(self.entity).filter(self.entity.fksucursal == idsurcusal).filter(
            self.entity.estado == True)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Domicilio.", fecha=fecha,tabla="domicilio", identificador=a.id)
        super().insert(b)
        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Domicilio.", fecha=fecha,tabla="domicilio", identificador=a.id)
        super().insert(b)
        return a

    def delete(self, id,estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.estado = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Domicilio.", fecha=fecha, tabla="domicilio", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        return x



    def filtrar(self, idcondominio,domicilio):
        if idcondominio != "0":
            objeto = self.db.query(self.entity).filter(self.entity.fkcondominio == idcondominio).filter(self.entity.tipo == domicilio).order_by(self.entity.nombre.asc()).all()

        else:
            objeto = self.db.query(self.entity).filter(self.entity.tipo == domicilio).order_by(self.entity.nombre.asc()).all()

        return objeto

    def importar_excel(self, cname,user,ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['COD_CONDOMINIO','COD_DOMICILIO','NUMERO','UBICACION','INTERNO','TIPO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    cod_condominio = row[indices['COD_CONDOMINIO']].value
                    cod_domicilio =  row[indices['COD_DOMICILIO']].value
                    numero = row[indices['NUMERO']].value
                    ubicacion = row[indices['UBICACION']].value
                    interno = row[indices['INTERNO']].value
                    tipo = row[indices['TIPO']].value

                    if cod_condominio is not None and cod_domicilio is not None:
                        query = self.db.query(self.entity).filter(
                            self.entity.codigo == str(cod_domicilio)).all()

                        cod_condominio = cod_condominio.replace(" ", "")

                        query_condominio = self.db.query(Condominio).filter(Condominio.codigo == str(cod_condominio)).first()

                        if query_condominio:
                            idcondominio = query_condominio.id
                            if not query:
                                cod_domicilio = cod_domicilio.replace(" ", "")
                                ubicacion = ubicacion.replace(" ", "")
                                domi = Domicilio(codigo=str(cod_domicilio),
                                                 numero=str(numero),
                                                 ubicacion=str(ubicacion),
                                                 interno=str(interno),
                                                 tipo=str(tipo),
                                                 fkcondominio=int(idcondominio))

                                self.db.merge(domi)
                                self.db.flush()
                            else:
                                print(str(i)+" "+ "No agregado")

                        else:
                            return {'message': 'Falta Codigo Condominio', 'success': False}
                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}