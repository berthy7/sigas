from .models import *
from ...common.managers import *
from sqlalchemy.exc import IntegrityError
from ...common.managers import *
from ...operaciones.bitacora.managers import *
from ...usuarios.usuario.managers import *
from ..residente.managers import *
from ..marca.managers import *
from ..modelo.managers import *
from ..nropase.managers import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font



class VehiculoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Vehiculo, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.placa.asc()).all()

    def listar_disponibles(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(and_(self.entity.fkresidente == None,self.entity.fkinvitado == None)).order_by(self.entity.placa.asc()).all()


    def obtener_x_id(self,idvehiculo):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == idvehiculo).first()

    def obtener_x_placa(self,placavehiculo):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.placa == placavehiculo).first()

    def consultar_vehiculo_invitado(self,idvehiculo,idinvitado):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == idvehiculo).filter(self.entity.fkinvitado == idinvitado).first()

    def consultar_vehiculo(self,idvehiculo,idresidente):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == idvehiculo).filter(self.entity.fkresidente == idresidente).first()

    def registrar_vehiculo(self, diccionario):
        placa = diccionario['placa']
        fktipo = diccionario['fktipo']
        fkmarca = diccionario['fkmarca']
        fkmodelo = diccionario['fkmodelo']
        fkcolor = diccionario['fkcolor']
        idinvitado = diccionario['fkinvitado']

        if fkmarca == "":
            fkmarca = None
        elif int(fkmarca) == 0:
            marc = Marca(nombre=diccionario['nombre_marca'])
            obj_marca = super().insert(marc)
            fkmarca = obj_marca.id
        if fkmodelo == "":
            fkmodelo = None

        vehiculo = VehiculoManager(self.db).obtener_x_placa(placa)

        if vehiculo:
            return vehiculo.id
        else:
            if placa == "":
                return None
            else:
                # dict_vehiculo = dict(placa=placa, tipo=tipo,marca=marca,color=color,fkpropietarioinvitado=idinvitado)
                dict_vehiculo = dict(placa=placa, fktipo=fktipo, fkmarca=fkmarca, fkmodelo=fkmodelo, fkcolor=fkcolor)
                objeto = VehiculoManager(self.db).entity(**dict_vehiculo)
                fecha = BitacoraManager(self.db).fecha_actual()
                a = super().insert(objeto)
                b = Bitacora(fkusuario=diccionario['user'], ip=diccionario['ip'], accion="Registro Vehiculo.", fecha=fecha,
                             tabla="vehiculo", identificador=a.id)
                super().insert(b)

        return a.id

    def registrar_vehiculo_invitado(self, diccionario,idinvitado):
        placa = diccionario['placa']
        fktipo = diccionario['fktipo']
        fkmarca = diccionario['fkmarca']
        fkmodelo = diccionario['fkmodelo']
        fkcolor = diccionario['fkcolor']

        if fkmarca == "":
            fkmarca = None
        elif int(fkmarca) == 0:
            marc = Marca(nombre=diccionario['nombre_marca'])
            obj_marca = super().insert(marc)
            fkmarca = obj_marca.id
        if fkmodelo == "":
            fkmodelo = None


        if diccionario['id'] !="":
            respuesta = VehiculoManager(self.db).consultar_vehiculo_invitado(diccionario['id'],idinvitado)
            if respuesta is None:
                dict_vehiculo = dict(placa=placa, fktipo=fktipo, fkmarca=fkmarca, fkmodelo=fkmodelo, fkcolor=fkcolor)
                objeto = VehiculoManager(self.db).entity(**dict_vehiculo)
                objeto.fkinvitado = idinvitado
                super().update(objeto)
        else:
            diccionario['id'] = None
            dict_vehiculo = dict(placa=placa, fktipo=fktipo, fkmarca=fkmarca, fkmodelo=fkmodelo, fkcolor=fkcolor)
            objeto = VehiculoManager(self.db).entity(**dict_vehiculo)
            objeto.fkinvitado = idinvitado
            super().insert(objeto)

    def registrar_vehiculo_residente(self, diccionario,idresidente):
        if diccionario['fkmodelo'] == "":
            diccionario['fkmodelo'] = None

        if diccionario['id'] !="":
            respuesta = VehiculoManager(self.db).consultar_vehiculo(diccionario['id'],idresidente)
            if respuesta is None:
                objeto = VehiculoManager(self.db).entity(**diccionario)
                objeto.fkresidente = idresidente
                a = super().update(objeto)
                if a.fknropase:
                    NropaseManager(self.db).situacion(a.fknropase, "Ocupado")
            elif respuesta.fknropase is None:
                respuesta.fknropase = diccionario['fknropase']
                a = super().update(respuesta)
                if a.fknropase:
                    NropaseManager(self.db).situacion(a.fknropase, "Ocupado")
            elif int(diccionario['fknropase']) != respuesta.fknropase:
                NropaseManager(self.db).situacion(respuesta.fknropase, "Libre")
                NropaseManager(self.db).situacion(diccionario['fknropase'], "Ocupado")


        else:
            diccionario['id'] = None
            objeto = VehiculoManager(self.db).entity(**diccionario)
            objeto.fkresidente = idresidente
            a = super().insert(objeto)

            if a.fknropase:
                NropaseManager(self.db).situacion(a.fknropase, "Ocupado")


    def insert(self, diccionary):
        objeto = VehiculoManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()
    
        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Vehiculo.", fecha=fecha, tabla="vehiculo",
                     identificador=a.id)
        super().insert(b)
        return a
    
    
    def update(self, diccionary):
        objeto = VehiculoManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()
    
        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Vehiculo.", fecha=fecha, tabla="vehiculo",
                     identificador=a.id)
        super().insert(b)
        return a
    
    
    def delete(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.estado = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Vehiculo.", fecha=fecha, tabla="vehiculo", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
    
        return x

    def obtener_vehiculo(self, placa,color,tipovehiculo,marca,modelo,tarjeta):

        query_vehiculo = self.db.query(Vehiculo).filter(Vehiculo.placa == str(placa)).first()
        query_tarjeta = self.db.query(Nropase).filter(Nropase.tarjeta == str(tarjeta)).first()

        if query_tarjeta:
            idtarjeta = query_tarjeta.id
        else:
            idtarjeta = None

        if query_vehiculo:

            nombre_modelo = ""
            if query_vehiculo.fkmodelo:
                nombre_modelo = query_vehiculo.modelo.nombre


            return  dict(id=query_vehiculo.id,placa="",fkcolor="",fktipo="",fkmarca="",nombre_marca=query_vehiculo.marca.nombre,fkmodelo="",nombre_modelo=nombre_modelo, fknropase=idtarjeta)
        else:
            if placa:

                query_tipo = self.db.query(Tipovehiculo).filter(Tipovehiculo.nombre == str(tipovehiculo)).first()
                if query_tipo:
                    idtipo = query_tipo.id
                else:
                    tip = Tipovehiculo(nombre=tipovehiculo)
                    obj_tipo = super().insert(tip)
                    idtipo = obj_tipo.id

                query_color = self.db.query(Color).filter(Color.nombre == str(color)).first()
                if query_color:
                    idcolor = query_color.id
                else:
                    col = Color(nombre=color)
                    obj_color = super().insert(col)
                    idcolor = obj_color.id

                query_marca = self.db.query(Marca).filter(Marca.nombre == str(marca)).first()
                if query_marca:
                    idmarca = query_marca.id
                else:
                    marc = Marca(nombre=marca)
                    obj_marca = super().insert(marc)
                    idmarca = obj_marca.id


                if modelo:
                    query_modelo = self.db.query(Modelo).filter(Modelo.nombre == str(modelo)).first()
                    if query_modelo:
                        idmodelo = query_modelo.id
                    else:
                        mod = Modelo(nombre=modelo,fkmarca=idmarca)
                        obj_modelo = super().insert(mod)
                        idmodelo = obj_modelo.id

                else:
                    idmodelo = modelo

                return  dict(id="", placa=placa, fkcolor=idcolor, fktipo=idtipo, fkmarca=idmarca,nombre_marca=marca, fkmodelo=idmodelo,nombre_modelo=modelo, fknropase=idtarjeta)
            else:
                return ""

    def vehiculo_excel(self, ):
        cname = "Vehiculos.xlsx"

        invitados = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'PLACA'
        ws['C' + str(indice)] = 'FKTIPO'
        ws['D' + str(indice)] = 'FKCOLOR'
        ws['E' + str(indice)] = 'MARCA'
        ws['F' + str(indice)] = 'FKMODELO'
        ws['G' + str(indice)] = 'FKRESIDENTE'
        ws['H' + str(indice)] = 'FKINVITADO'
        ws['I' + str(indice)] = 'TARJETA'
        ws['J' + str(indice)] = 'ESTADO'

        for i in invitados:
            indice = indice + 1
            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.placa
            ws['C' + str(indice)] = i.fktipo
            ws['D' + str(indice)] = i.fkcolor
            ws['E' + str(indice)] = i.marca.nombre
            ws['F' + str(indice)] = i.fkmodelo
            ws['G' + str(indice)] = i.fkresidente
            ws['H' + str(indice)] = i.fkinvitado
            ws['I' + str(indice)] = i.nropase.tarjeta if i.fknropase is not None else None
            ws['J' + str(indice)] = i.estado


        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'PLACA', 'FKTIPO', 'FKCOLOR', 'MARCA', 'FKMODELO', 'FKRESIDENTE', 'FKINVITADO',
                        'TARJETA', 'ESTADO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    id = row[indices['ID']].value
                    placa = row[indices['PLACA']].value
                    fktipo = row[indices['FKTIPO']].value
                    fkcolor = row[indices['FKCOLOR']].value
                    marca = row[indices['MARCA']].value
                    fkmodelo = row[indices['FKMODELO']].value
                    fkresidente = row[indices['FKRESIDENTE']].value
                    fkinvitado = row[indices['FKINVITADO']].value
                    tarjeta = row[indices['TARJETA']].value
                    estado = row[indices['ESTADO']].value

                    if id is not None:
                        idmarca = ""
                        idtarjeta = ""

                        if fkresidente is None:
                            query = self.db.query(Vehiculo).filter(Vehiculo.placa == placa).first()

                            if not query:


                                query_marca = self.db.query(Marca).filter(Marca.nombre == str(marca)).first()
                                if query_marca:
                                    idmarca = query_marca.id
                                else:
                                    marc = Marca(nombre=marca)
                                    obj_marca = super().insert(marc)
                                    idmarca = obj_marca.id

                                query_tarjeta = self.db.query(Nropase).filter(Nropase.tarjeta == str(tarjeta)).first()

                                if query_tarjeta:
                                    idtarjeta = query_tarjeta.id
                                else:
                                    idtarjeta = None

                                invi = Vehiculo(placa=placa, fktipo=fktipo, fkcolor=fkcolor,
                                                fkmarca=idmarca, fkmodelo=fkmodelo, fkresidente=fkresidente, fkinvitado=None,
                                                fknropase=idtarjeta, estado=estado)

                                self.db.merge(invi)
                                self.db.flush()

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

class TipovehiculoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Tipovehiculo, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()


class ColorManager(SuperManager):
    def __init__(self, db):
        super().__init__(Color, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).all()
