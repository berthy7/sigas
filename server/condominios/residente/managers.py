from .models import *
from ...usuarios.usuario.managers import *
from sqlalchemy.exc import IntegrityError
from ...common.managers import *
from ...operaciones.bitacora.managers import *
from ..domicilio.managers import *
from ..condominio.models import *
from ..vehiculo.managers import *
from ..nropase.managers import *
from ...dispositivos.dispositivo.managers import *
from datetime import datetime, timedelta, time, date

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


class ResidenteManager(SuperManager):
    def __init__(self, db):
        super().__init__(Residente, db)


    def get_all(self):
        return self.db.query(self.entity)


    def obtener_x_id(self, id):
        return self.db.query(self.entity).filter(self.entity.id == id).first()

    def obtener_x_codigo(self, codigo):
        return self.db.query(self.entity).filter(self.entity.codigo == codigo).first()

    def get_all_by_lastname(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.apellidop.asc()).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.apellidop.asc()).all()


    def insert(self, diccionary):

        estado = diccionary['acceso'][0]['estado']
        if diccionary['b_fknropase'] == "0":
            if diccionary['fknropase']:
                NropaseManager(self.db).situacion(diccionary['fknropase'], "Libre")
                diccionary['fknropase'] = None

        if diccionary['fknropase'] == "":
            diccionary['fknropase'] = None

        objeto = ResidenteManager(self.db).entity(**diccionary)
        if objeto.fechanacimiento != "":
            objeto.fechanacimiento = datetime.strptime(objeto.fechanacimiento, '%d/%m/%Y')
        else:
            objeto.fechanacimiento = None

        for acce in objeto.acceso:
            acce.fechai = datetime.strptime(acce.fechai, '%d/%m/%Y')
            acce.fechaf = datetime.strptime(acce.fechaf, '%d/%m/%Y')

        fecha = BitacoraManager(self.db).fecha_actual()
        objeto.vehiculos = []
        objeto.estado = estado

        if objeto.codigo == "":
            objeto.codigo = None

        a = super().insert(objeto)
        # a = self.db.query(self.entity).filter(self.entity.id == 14).first()

        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Residente.", fecha=fecha,tabla="residente", identificador=a.id)
        super().insert(b)
        print("codigo residente : " + str(a.codigo))
        if a.codigo == None:
            print("codigo none")

            a.codigo = a.id

            inicial_cod = 100000
            a.codigoqr= a.id + inicial_cod
            super().update(a)

        if a.fknropase:
            NropaseManager(self.db).situacion(a.fknropase, "Ocupado")

        for vehi in diccionary['vehiculos']:
            VehiculoManager(self.db).registrar_vehiculo_residente(vehi,a.id)

        idcondominio=None

        for domi in a.domicilios:
            if domi.vivienda:
                d = domi.fkdomicilio
                idcondominio = DomicilioManager(self.db).obtener_fkcondominio(d)
                break


        if a.fknropase:
            tarjeta = a.nropase.tarjeta

        else:
            tarjeta = ""


        password = UsuarioManager(self.db).generar_contraseña()
        dict_usuario = dict(nombre=a.nombre,apellidop=a.apellidop,apellidom=a.apellidom,ci=a.ci,expendido=a.expendido,
                            correo=a.correo,telefono=a.telefono,username=a.correo,password=password,default=password,
                            fkrol=7,fkresidente=a.id,codigoqr_residente=a.codigoqr,tarjeta_residente=tarjeta, fkcondominio=idcondominio,sigas=False,
                            user_id=objeto.user,ip=objeto.ip,estado=estado,enabled=True)

        return dict_usuario

    def update(self, diccionary):
        estado = diccionary['acceso'][0]['estado']

        if diccionary['b_fknropase'] == "0":
            if diccionary['fknropase']:
                NropaseManager(self.db).situacion(diccionary['fknropase'], "Libre")
                diccionary['fknropase'] = None
            else:
                diccionary['fknropase'] = None

        if diccionary['fknropase'] == "":
            diccionary['fknropase'] = None

        if diccionary['actual_fknropase'] != diccionary['fknropase']:
            if diccionary['actual_fknropase'] != "":
                NropaseManager(self.db).situacion(diccionary['actual_fknropase'], "Libre")


        for vehi in diccionary['vehiculos']:
            vehiculo = VehiculoManager(self.db).registrar_vehiculo_residente(vehi,diccionary['id'])
            if vehiculo:
                vehi['id'] = vehiculo.id

        objeto = ResidenteManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        if objeto.fechanacimiento == "":
            objeto.fechanacimiento = None
        else:
            objeto.fechanacimiento = datetime.strptime(objeto.fechanacimiento, '%d/%m/%Y')

        for acce in objeto.acceso:
            acce.fechai = datetime.strptime(acce.fechai, '%d/%m/%Y')
            acce.fechaf = datetime.strptime(acce.fechaf, '%d/%m/%Y')

        objeto.estado = estado

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Residente.", fecha=fecha,tabla="residente", identificador=a.id)
        super().insert(b)

        if a.fknropase:
            NropaseManager(self.db).situacion(a.fknropase, "Ocupado")

        use = self.db.query(Usuario).filter(Usuario.fkresidente == a.id).first()
        use.enabled = estado
        self.db.merge(use)
        self.db.commit()


        return a

    def delete(self, dic):
        if dic['fkresidente_change'] == "":
            dic['fkresidente_change'] = None

        if dic['codigo'] == "":
            dic['codigo'] = None

        x = self.db.query(Residente).filter(Residente.codigo == dic['codigo']).first()
        x_change = self.db.query(Residente).filter(Residente.codigo == dic['fkresidente_change']).first()

        xacceso = self.db.query(ResidenteAcceso).filter(ResidenteAcceso.fkresidente == x.id).first()

        x.estado = False
        xacceso.estado = False
        x.enabled = False

        usuario = self.db.query(Usuario).filter(Usuario.fkresidente == x.id).first()

        if usuario:

            usuario.estado =  False
            usuario.enabled = False
            self.db.merge(usuario)

        if x_change:

            for domi in x.domicilios:

                residenteDomicilio = self.db.query(ResidenteDomicilio).filter(ResidenteDomicilio.fkresidente == x_change.id) \
                    .filter(ResidenteDomicilio.fkdomicilio == domi.fkdomicilio).first()

                if residenteDomicilio is None:
                    super().insert(ResidenteDomicilio(fkresidente=x_change.id,fkdomicilio=domi.fkdomicilio))



        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=dic['user'], ip=dic['ip'], accion= "Elimino Residente", fecha=fecha, tabla="residente", identificador=x.id)
        super().insert(b)
        self.db.merge(x)

        self.db.merge(xacceso)
        self.db.commit()

        principal = self.db.query(Principal).first()

        if principal.estado:

            try:
                if x.fkcondominio:

                    if x.condominio.ip_publica != "":

                        url = "http://" + x.condominio.ip_publica + ":" + x.condominio.puerto + "/api/v1/sincronizar_residente_deshabilitar"

                        headers = {'Content-Type': 'application/json'}
                        string = dic
                        cadena = json.dumps(string)
                        body = cadena
                        resp = requests.post(url, data=body, headers=headers, verify=False)
                        response = json.loads(resp.text)
                        print("Respuesta Sincro baja de residente")
                        print(response)


            except Exception as e:
                # Other errors are possible, such as IOError.
                print("Error de conexion: " + str(e))

        return "Elimino Residente"

    # def importar_excel(self, cname,user,ip):
    #     i = 0
    #     try:
    #         wb = load_workbook(filename="server/common/resources/uploads/" + cname)
    #         ws = wb.active
    #         colnames = ['NOMBRE','APELLIDOP','APELLIDOM', 'CI', 'UBICACION', 'SEXO', 'TELEFONO', 'TIPO', 'CORREO',
    #                     'DOMICILIO','NUMERO','TARTEJAPEATONAL', 'PLACA', 'COLOR', 'TIPOVEHICULO', 'MARCA', 'MODELO', 'TARJETAVEHICULAR', 'FECHAI', 'FECHAF']
    #         indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
    #                    cell[0].value in colnames}
    #         if len(indices) == len(colnames):
    #             for row in ws.iter_rows(min_row=2):
    #                 nombre =  row[indices['NOMBRE']].value
    #                 apellidop = row[indices['APELLIDOP']].value
    #
    #                 apellidom = row[indices['APELLIDOM']].value if row[indices['APELLIDOM']].value else ""
    #
    #                 ci = row[indices['CI']].value
    #                 ubicacion = row[indices['UBICACION']].value
    #                 sexo = row[indices['SEXO']].value
    #                 telefono =  row[indices['TELEFONO']].value
    #                 tipo =  row[indices['TIPO']].value
    #                 correo = row[indices['CORREO']].value
    #                 domicilio =  row[indices['DOMICILIO']].value
    #                 numero = row[indices['NUMERO']].value
    #                 tarjetapeatonal = row[indices['TARTEJAPEATONAL']].value
    #                 placa = row[indices['PLACA']].value
    #                 color = row[indices['COLOR']].value
    #                 tipovehiculo = row[indices['TIPOVEHICULO']].value
    #                 marca =  row[indices['MARCA']].value
    #                 modelo =  row[indices['MODELO']].value
    #                 tarjetavehicular = row[indices['TARJETAVEHICULAR']].value
    #                 fechai =  row[indices['FECHAI']].value
    #                 fechaf = row[indices['FECHAF']].value
    #
    #                 fechanacimiento = None
    #                 fechai = fechai.strftime('%d/%m/%Y')
    #                 fechaf = fechaf.strftime('%d/%m/%Y')
    #
    #
    #                 if nombre is not None and correo is not None and domicilio is not None and numero is not None:
    #
    #                     query = self.db.query(self.entity).filter(
    #                         self.entity.correo == str(correo)).first()
    #
    #                     list_domicilio = list()
    #                     list_vehiculo = list()
    #                     list_acceso = list()
    #
    #                     if apellidom is None:
    #                         apellidom = ""
    #
    #                     if telefono is None:
    #                         telefono = ""
    #
    #                     # domicilio = domicilio.replace(" ", "")
    #
    #                     query_domicilio = self.db.query(Domicilio).filter(Domicilio.ubicacion == str(domicilio)).filter(Domicilio.numero == str(numero)).first()
    #
    #                     if query_domicilio:
    #                         if not query:
    #
    #                             dict_vehiculo = VehiculoManager(self.db).obtener_vehiculo(placa, color, tipovehiculo,marca, modelo,tarjetavehicular)
    #                             list_domicilio.append(dict(fkdomicilio=query_domicilio.id,codigo_domicilio=query_domicilio.codigo,vivienda=True))
    #                             if dict_vehiculo != "":
    #                                 list_vehiculo.append(dict_vehiculo)
    #
    #                             query_tarjeta = self.db.query(Nropase).filter(Nropase.tarjeta == str(tarjetapeatonal)).first()
    #
    #                             if query_tarjeta:
    #                                 idtarjeta = query_tarjeta.id
    #                             else:
    #                                 idtarjeta = None
    #
    #                             list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
    #                             residente = dict( nombre=str(nombre),
    #                                               apellidop=str(apellidop),
    #                                               apellidom=str(apellidom),
    #                                               ci=str(ci),
    #                                               expendido=str(ubicacion),
    #                                               sexo=str(sexo),
    #                                               fechanacimiento=fechanacimiento,
    #                                               telefono=str(telefono),
    #                                               tipo=str(tipo),
    #                                               correo=str(correo),
    #                                               fknropase=idtarjeta,
    #                                               b_fknropase=idtarjeta,
    #                                               domicilios= list_domicilio,
    #                                               vehiculos= list_vehiculo,
    #                                               acceso=list_acceso,
    #                                               user=user,
    #                                               ip=ip)
    #
    #                             dict_usuario = ResidenteManager(self.db).insert(residente)
    #
    #                             c = UsuarioManager(self.db).insert_residente(dict_usuario)
    #
    #                             principal = self.db.query(Principal).first()
    #
    #                             if principal.estado:
    #
    #                                 if c['response'].condominio.ip_publica != "":
    #                                     url = "http://" + c['response'].condominio.ip_publica + ":" + c['response'].condominio.puerto + "/api/v1/sincronizar_residente"
    #
    #                                     headers = {'Content-Type': 'application/json'}
    #                                     dict_usuario['codigo'] = c['response'].id
    #
    #
    #                                     diccionary = dict(dict_usuario=dict_usuario,dict_residente=residente)
    #
    #                                     cadena = json.dumps(diccionary)
    #                                     body = cadena
    #                                     resp = requests.post(url, data=body, headers=headers, verify=False)
    #                                     response = json.loads(resp.text)
    #
    #                                     print(response)
    #
    #                             # self.db.merge(propietario)
    #                             # self.db.flush()
    #                 else:
    #
    #                     self.db.rollback()
    #                     return {'message': 'Hay Columnas vacias', 'success': False}
    #
    #             self.db.commit()
    #             return {'message': 'Importado Todos Correctamente.', 'success': True}
    #         else:
    #             return {'message': 'Columnas Faltantes', 'success': False}
    #     except IntegrityError as e:
    #         self.db.rollback()
    #         if 'UNIQUE constraint failed: residente.ci' in str(e):
    #             return {'message': 'CI duplicado', 'success': False}
    #         if 'UNIQUE constraint failed: residente.codigo' in str(e):
    #             return {'message': 'codigo de residente', 'success': False}
    #         return {'message': str(e), 'success': False}

    def importar_excel(self, cname, user, ip):
        i = 0
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['Apellido_Paterno', 'Apellido_Materno', 'Nombre', 'Tipo_de_Residente', 'Interno', 'Nro_de_Casa', 'Calle', 'Celular', 'Correo',
                        'Ci','PLACA', 'COLOR', 'TIPOVEHICULO', 'MARCA', 'MODELO',
                        'TARJETAVEHICULAR', 'FECHAI', 'FECHAF']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    nombre = row[indices['Nombre']].value
                    apellidop = row[indices['Apellido_Paterno']].value if row[indices['Apellido_Paterno']].value else " "

                    apellidom = row[indices['Apellido_Materno']].value if row[indices['Apellido_Materno']].value else " "

                    ci = row[indices['Ci']].value
                    ubicacion = ''
                    sexo = ''
                    telefono = ''
                    tipo = row[indices['Tipo_de_Residente']].value
                    correo = row[indices['Correo']].value
                    domicilio = row[indices['Calle']].value
                    numero = row[indices['Nro_de_Casa']].value
                    tarjetapeatonal = ''
                    placa = row[indices['PLACA']].value
                    color = row[indices['COLOR']].value
                    tipovehiculo = row[indices['TIPOVEHICULO']].value
                    marca = row[indices['MARCA']].value
                    modelo = row[indices['MODELO']].value
                    tarjetavehicular = row[indices['TARJETAVEHICULAR']].value
                    fechai = row[indices['FECHAI']].value
                    fechaf = row[indices['FECHAF']].value

                    print(nombre)

                    fechanacimiento = None
                    fechai = fechai.strftime('%d/%m/%Y')
                    fechaf = fechaf.strftime('%d/%m/%Y')

                    if nombre:
                        if correo:
                            query = self.db.query(self.entity).filter(
                                self.entity.correo == str(correo)).first()
                        else:
                            query = self.db.query(self.entity).filter(
                                self.entity.nombre == str(nombre)).filter(
                                self.entity.apellidop == str(apellidop)).filter(
                                self.entity.apellidom == str(apellidom)).first()


                        list_domicilio = list()
                        list_vehiculo = list()
                        list_acceso = list()

                        if apellidom is None:
                            apellidom = ""

                        if telefono is None:
                            telefono = ""

                        # domicilio = domicilio.replace(" ", "")

                        query_domicilio = self.db.query(Domicilio).filter(Domicilio.ubicacion == str(domicilio)).filter(
                            Domicilio.numero == str(numero)).first()

                        if query_domicilio:
                            if not query:

                                dict_vehiculo = VehiculoManager(self.db).obtener_vehiculo(placa, color, tipovehiculo,
                                                                                          marca, modelo,
                                                                                          tarjetavehicular)
                                list_domicilio.append(
                                    dict(fkdomicilio=query_domicilio.id, codigo_domicilio=query_domicilio.codigo,
                                         vivienda=True))
                                if dict_vehiculo != "":
                                    list_vehiculo.append(dict_vehiculo)

                                query_tarjeta = self.db.query(Nropase).filter(
                                    Nropase.tarjeta == str(tarjetapeatonal)).first()

                                if query_tarjeta:
                                    idtarjeta = query_tarjeta.id
                                else:
                                    idtarjeta = None

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(nombre),
                                                 apellidop=str(apellidop),
                                                 apellidom=str(apellidom),
                                                 ci=str(ci),
                                                 expendido=str(ubicacion),
                                                 sexo=str(sexo),
                                                 fechanacimiento=fechanacimiento,
                                                 telefono=str(telefono),
                                                 tipo=str(tipo),
                                                 correo=str(correo),
                                                 fknropase=idtarjeta,
                                                 b_fknropase=idtarjeta,
                                                 domicilios=list_domicilio,
                                                 vehiculos=list_vehiculo,
                                                 acceso=list_acceso,
                                                 user=user,
                                                 ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)

                                c = UsuarioManager(self.db).insert_residente(dict_usuario)

                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()

                            else:
                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == query.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(
                                        ResidenteDomicilio(fkresidente=query.id, fkdomicilio=query_domicilio.id))

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: residente.ci' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: residente.codigo' in str(e):
                return {'message': 'codigo de residente', 'success': False}
            return {'message': str(e), 'success': False}

    def importar_excel_nuevo(self, cname, user, ip):
        i = 0
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['CODIGO', 'CALLE', 'INTERNO', 'COPROPIETARIO_1', 'COPROPIETARIO_2', 'COPROPIETARIO_3', 'INQUILINO_1', 'INQUILINO_2',
                        'CORREO', 'Telf_Propietario_1', 'Telf_Propietario_2', 'Telf_Propietario_3', 'Telf_Inquilino_1', 'Telf_Inquilino_2']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    codigo = row[indices['CODIGO']].value
                    calle = row[indices['CALLE']].value
                    interno = row[indices['INTERNO']].value
                    copropietario1 = row[indices['COPROPIETARIO_1']].value
                    copropietario2 = row[indices['COPROPIETARIO_2']].value
                    copropietario3 = row[indices['COPROPIETARIO_3']].value
                    inquilino1 = row[indices['INQUILINO_1']].value
                    inquilino2 = row[indices['INQUILINO_2']].value
                    correo = row[indices['CORREO']].value
                    telfPropietario1 = row[indices['Telf_Propietario_1']].value
                    telfPropietario2 = row[indices['Telf_Propietario_2']].value
                    telfPropietario3 = row[indices['Telf_Propietario_3']].value
                    telfInquilino1 = row[indices['Telf_Inquilino_1']].value
                    telfInquilino2 = row[indices['Telf_Inquilino_2']].value

                    list_domicilio = list()
                    list_vehiculo = list()
                    list_acceso = list()

                    fechai = BitacoraManager(self.db).fecha_actual()
                    fechaf = fechai+ timedelta(days=365)

                    fechai =fechai.strftime('%d/%m/%Y')

                    fechaf = fechaf.strftime('%d/%m/%Y')

                    query_domicilio = self.db.query(Domicilio).filter(Domicilio.ubicacion == calle).filter(
                        Domicilio.numero == str(codigo)).first()

                    if query_domicilio:

                        # Copropietario 1

                        if copropietario1 == "Hugo Noya Gutierrez":
                            print("si")

                        query = self.db.query(self.entity).filter(
                            self.entity.correo == str(correo)).first()

                        if query:

                            query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                ResidenteDomicilio.fkresidente == query.id).filter(
                                ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                            if not query_residentedomicilio:

                                super().insert(ResidenteDomicilio(fkresidente=query.id,fkdomicilio=query_domicilio.id))

                        else:

                            query_nombre = self.db.query(self.entity).filter(
                                self.entity.nombre == str(copropietario1)).first()

                            if query_nombre:

                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == query_nombre.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(ResidenteDomicilio(fkresidente=query_nombre.id,
                                                                      fkdomicilio=query_domicilio.id))
                            else:

                                list_domicilio.append(dict(fkdomicilio=query_domicilio.id, codigo_domicilio=query_domicilio.codigo,
                                         vivienda=True))

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(copropietario1),apellidop="",apellidom="",ci="",expendido="",
                                                 sexo="",fechanacimiento=None,telefono=str(telfPropietario1),
                                                 tipo="Copropietario",correo=str(correo),fknropase=None,
                                                 b_fknropase=None,domicilios=list_domicilio,vehiculos=list_vehiculo,
                                                 acceso=list_acceso,user=user,ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)


                                if dict_usuario['correo'] == 'None':
                                    dict_usuario['correo'] = str(dict_usuario['fkresidente'])
                                    dict_usuario['username'] = str(dict_usuario['fkresidente'])

                                    residente_query = self.db.query(Residente).filter(
                                        Residente.id == dict_usuario['fkresidente']).first()

                                    residente_query.correo = residente_query.id

                                    self.db.merge(residente_query)
                                    self.db.commit()


                                c = UsuarioManager(self.db).insert_residente(dict_usuario)

                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()


                        # //////////////////////////////////////////

                        # Copropietario 2

                        if copropietario2:

                            query2 = self.db.query(self.entity).filter(
                                self.entity.nombre == str(copropietario2)).first()

                            if query2:

                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == query2.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(ResidenteDomicilio(fkresidente=query2.id,
                                                                      fkdomicilio=query_domicilio.id))

                            else:

                                list_domicilio.append(dict(fkdomicilio=query_domicilio.id,
                                                           codigo_domicilio=query_domicilio.codigo,
                                                           vivienda=True))

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(copropietario2), apellidop="", apellidom="", ci="",
                                                 expendido="",
                                                 sexo="", fechanacimiento=None, telefono=str(telfPropietario2),
                                                 tipo="Copropietario", correo=None, fknropase=None,
                                                 b_fknropase=None, domicilios=list_domicilio,
                                                 vehiculos=list_vehiculo,
                                                 acceso=list_acceso, user=user, ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)

                                dict_usuario['correo'] = str(dict_usuario['fkresidente'])
                                dict_usuario['username'] = str(dict_usuario['fkresidente'])

                                residente_query = self.db.query(Residente).filter(
                                    Residente.id == dict_usuario['fkresidente']).first()

                                residente_query.correo = residente_query.id

                                self.db.merge(residente_query)
                                self.db.commit()

                                c = UsuarioManager(self.db).insert_residente(dict_usuario)


                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()


                        # //////////////////////////////////////////


                        # Copropietario 3

                        if copropietario3:

                            query3 = self.db.query(self.entity).filter(
                                self.entity.nombre == str(copropietario3)).first()

                            if query3:

                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == query3.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(ResidenteDomicilio(fkresidente=query3.id,
                                                                      fkdomicilio=query_domicilio.id))

                            else:

                                list_domicilio.append(dict(fkdomicilio=query_domicilio.id,
                                                           codigo_domicilio=query_domicilio.codigo,
                                                           vivienda=True))

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(copropietario3), apellidop="", apellidom="", ci="",
                                                 expendido="",
                                                 sexo="", fechanacimiento=None, telefono=str(telfPropietario3),
                                                 tipo="Copropietario", correo=None, fknropase=None,
                                                 b_fknropase=None, domicilios=list_domicilio,
                                                 vehiculos=list_vehiculo,
                                                 acceso=list_acceso, user=user, ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)

                                dict_usuario['correo'] = str(dict_usuario['fkresidente'])
                                dict_usuario['username'] = str(dict_usuario['fkresidente'])

                                residente_query = self.db.query(Residente).filter(
                                    Residente.id == dict_usuario['fkresidente']).first()

                                residente_query.correo = residente_query.id

                                self.db.merge(residente_query)
                                self.db.commit()


                                c = UsuarioManager(self.db).insert_residente(dict_usuario)

                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()


                        # //////////////////////////////////////////


                        # inquilino1

                        if inquilino1:

                            queryinquilino1 = self.db.query(self.entity).filter(
                                self.entity.nombre == str(inquilino1)).first()

                            if queryinquilino1:

                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == queryinquilino1.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(ResidenteDomicilio(fkresidente=queryinquilino1.id,
                                                                      fkdomicilio=query_domicilio.id))

                            else:

                                list_domicilio.append(dict(fkdomicilio=query_domicilio.id,
                                                           codigo_domicilio=query_domicilio.codigo,
                                                           vivienda=True))

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(inquilino1), apellidop="", apellidom="", ci="",
                                                 expendido="",
                                                 sexo="", fechanacimiento=None, telefono=str(telfInquilino1),
                                                 tipo="Inquilino", correo=None, fknropase=None,
                                                 b_fknropase=None, domicilios=list_domicilio,
                                                 vehiculos=list_vehiculo,
                                                 acceso=list_acceso, user=user, ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)

                                dict_usuario['correo'] = str(dict_usuario['fkresidente'])
                                dict_usuario['username'] = str(dict_usuario['fkresidente'])

                                residente_query = self.db.query(Residente).filter(
                                    Residente.id == dict_usuario['fkresidente']).first()

                                residente_query.correo = residente_query.id

                                self.db.merge(residente_query)
                                self.db.commit()


                                c = UsuarioManager(self.db).insert_residente(dict_usuario)

                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()

                        # //////////////////////////////////////////

                        # inquilino2

                        if inquilino2:

                            queryinquilino2 = self.db.query(self.entity).filter(
                                self.entity.nombre == str(inquilino2)).first()

                            if queryinquilino2:

                                query_residentedomicilio = self.db.query(ResidenteDomicilio).filter(
                                    ResidenteDomicilio.fkresidente == queryinquilino2.id).filter(
                                    ResidenteDomicilio.fkdomicilio == query_domicilio.id).first()

                                if not query_residentedomicilio:
                                    super().insert(ResidenteDomicilio(fkresidente=queryinquilino2.id,
                                                                      fkdomicilio=query_domicilio.id))

                            else:

                                list_domicilio.append(dict(fkdomicilio=query_domicilio.id,
                                                           codigo_domicilio=query_domicilio.codigo,
                                                           vivienda=True))

                                list_acceso.append(dict(fechai=fechai, fechaf=fechaf, estado=False))
                                residente = dict(nombre=str(inquilino2), apellidop="", apellidom="", ci="",
                                                 expendido="",
                                                 sexo="", fechanacimiento=None, telefono=str(telfInquilino2),
                                                 tipo="Inquilino", correo=None, fknropase=None,
                                                 b_fknropase=None, domicilios=list_domicilio,
                                                 vehiculos=list_vehiculo,
                                                 acceso=list_acceso, user=user, ip=ip)

                                dict_usuario = ResidenteManager(self.db).insert(residente)

                                dict_usuario['correo'] = str(dict_usuario['fkresidente'])
                                dict_usuario['username'] = str(dict_usuario['fkresidente'])

                                residente_query = self.db.query(Residente).filter(
                                    Residente.id == dict_usuario['fkresidente']).first()

                                residente_query.correo = residente_query.id

                                self.db.merge(residente_query)
                                self.db.commit()

                                c = UsuarioManager(self.db).insert_residente(dict_usuario)

                                principal = self.db.query(Principal).first()

                                if principal.estado:

                                    if c['response'].condominio.ip_publica != "":
                                        url = "http://" + c['response'].condominio.ip_publica + ":" + c[
                                            'response'].condominio.puerto + "/api/v1/sincronizar_residente"

                                        headers = {'Content-Type': 'application/json'}
                                        dict_usuario['codigo'] = c['response'].id

                                        diccionary = dict(dict_usuario=dict_usuario, dict_residente=residente)

                                        cadena = json.dumps(diccionary)
                                        body = cadena
                                        resp = requests.post(url, data=body, headers=headers, verify=False)
                                        response = json.loads(resp.text)

                                        print(response)

                                        # self.db.merge(propietario)
                                        # self.db.flush()


                                    # //////////////////////////////////////////





                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: residente.ci' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: residente.codigo' in str(e):
                return {'message': 'codigo de residente', 'success': False}
            return {'message': str(e), 'success': False}

    def persona_excel(self, empleados):
        fecha = datetime.now()
        cname = "Empleados" + fecha.strftime('%Y-%m-%d') + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Empleados'

        indice = 0
        # --------------------------------------------------------------------
        indice = indice + 1
        ws['A' + str(indice)] = 'Nº'
        ws['B' + str(indice)] = 'PAIS'
        ws['C' + str(indice)] = 'DEPARTAMENTO'
        ws['D' + str(indice)] = 'CIUDAD'
        ws['E' + str(indice)] = 'EMPRESA'
        ws['F' + str(indice)] = 'SUCURSAL'
        ws['G' + str(indice)] = 'CENTRO DE COSTO'
        ws['H' + str(indice)] = 'SECCION'
        ws['I' + str(indice)] = 'CODIGO'
        ws['J' + str(indice)] = 'NOMBRE'
        ws['K' + str(indice)] = 'APELLIDO_P'
        ws['L' + str(indice)] = 'APELLIDO_M'
        ws['M' + str(indice)] = 'CI'
        ws['N' + str(indice)] = 'SEXO'
        ws['O' + str(indice)] = 'FECHA_NACIMIENTO'
        ws['P' + str(indice)] = 'CARGO'
        ws['Q' + str(indice)] = 'CONTRATO_PLAZO'
        ws['R' + str(indice)] = 'FECHA_INGRESO_NOM'
        ws['S' + str(indice)] = 'FECHA_VENCIMIENTO'
        ws['T' + str(indice)] = 'CORREO'

        ws['A' + str(indice)].font = Font(bold=True)
        ws['B' + str(indice)].font = Font(bold=True)
        ws['C' + str(indice)].font = Font(bold=True)
        ws['D' + str(indice)].font = Font(bold=True)
        ws['E' + str(indice)].font = Font(bold=True)
        ws['F' + str(indice)].font = Font(bold=True)
        ws['G' + str(indice)].font = Font(bold=True)
        ws['H' + str(indice)].font = Font(bold=True)
        ws['I' + str(indice)].font = Font(bold=True)
        ws['J' + str(indice)].font = Font(bold=True)
        ws['K' + str(indice)].font = Font(bold=True)
        ws['L' + str(indice)].font = Font(bold=True)
        ws['M' + str(indice)].font = Font(bold=True)
        ws['N' + str(indice)].font = Font(bold=True)
        ws['O' + str(indice)].font = Font(bold=True)
        ws['P' + str(indice)].font = Font(bold=True)
        ws['Q' + str(indice)].font = Font(bold=True)
        ws['R' + str(indice)].font = Font(bold=True)
        ws['S' + str(indice)].font = Font(bold=True)
        ws['T' + str(indice)].font = Font(bold=True)

        for i in empleados:
            x = ResidenteManager(self.db).get_dataemp(i['id'])

            indice = indice + 1
            ws['A' + str(indice)] = x[0].id
            ws['B' + str(indice)] = x[0].empleado[0].pais.nombre
            ws['C' + str(indice)] = x[0].empleado[0].domicilios.nombre
            ws['D' + str(indice)] = x[0].empleado[0].ciudad.nombre
            ws['E' + str(indice)] = x[0].empleado[0].sucursal.empresa.nombre
            ws['F' + str(indice)] = x[0].empleado[0].sucursal.nombre
            ws['G' + str(indice)] = x[0].empleado[0].centro.nombre
            ws['H' + str(indice)] = x[0].empleado[0].gerencia.nombre
            ws['I' + str(indice)] = x[0].empleado[0].codigo
            ws['J' + str(indice)] = x[0].nombres
            ws['K' + str(indice)] = x[0].apellidopaterno
            ws['L' + str(indice)] = x[0].apellidomaterno
            ws['M' + str(indice)] = x[0].ci
            ws['N' + str(indice)] = x[0].sexo
            ws['O' + str(indice)] = x[0].fechanacimiento.strftime('%d/%m/%Y')
            ws['P' + str(indice)] = x[0].empleado[0].cargo.nombre
            ws['Q' + str(indice)] = x[0].contrato[0].tipo
            ws['R' + str(indice)] = x[0].contrato[0].fechaIngreso.strftime('%d/%m/%Y')
            ws['S' + str(indice)] = x[0].contrato[0].fechaFin.strftime('%d/%m/%Y')
            ws['T' + str(indice)] = x[0].empleado[0].email

        for col in ws.columns:
            row_idx = 0
            max_length = 0
            column = col[0].column
            for cell in col:
                row_idx = row_idx + 1
                try:
                    max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            if column not in ['N']:
                ws.column_dimensions[column].width = adjusted_width

        wb.save("server/common/resources/downloads/residente/" + cname)
        return cname

    def get_all_data(self, id):
        list = {}
        c = 0
        for dataper in self.db.query(Residente).filter(Residente.estado == True).filter(Residente.id == id):
            #list[c] = dict(id=dataper.ValorParametro.id, nombre=dataper.ValorParametro.nombre, valor=dataper.ValorParametro.valor)
            c = c + 1
        return list

    def obtener_domicilios(self, idresidente):

        x = self.db.query(Domicilio).join(ResidenteDomicilio).filter(ResidenteDomicilio.fkresidente == idresidente).filter(ResidenteDomicilio.vivienda == True).filter(Domicilio.estado == True).order_by(Domicilio.tipo.desc(),Domicilio.ubicacion.asc()).first()

        return x


    def listar_residentes(self,usuario):

        if usuario.sigas:
            return self.db.query(Residente).filter(Residente.enabled).order_by(Residente.nombre.asc()).all()

        elif usuario.rol.nombre == "RESIDENTE":
            return self.db.query(Residente).filter(Residente.enabled).filter(Residente.id == usuario.fkresidente).order_by(Residente.nombre.asc()).all()

        else:
            return self.db.query(Residente).filter(Residente.enabled).filter(Residente.fkcondominio == usuario.fkcondominio).order_by(Residente.nombre.asc()).all()


    def validar_codigo(self, codigoautorizacion):
        domicilio = ""

        x = self.db.query(Residente).filter(Residente.estado == True).filter(Residente.codigoqr == codigoautorizacion).first()
        if x:
            for domi in x.domicilios:
                if domi.vivienda:
                    domicilio = domi.fkdomicilio
                    break

            residente = dict(idresidente= x.id,iddomicilio=domicilio,fotoresidente=x.foto,tipodocumento=4)
        else:
            residente = None
        return residente


    def actualizar_foto(self, data):
        try:
            print("servicio foto")
            print(str(data['user']))
            x = self.db.query(Usuario).filter(Usuario.id == data['user']).first()
            persona = self.db.query(Residente).filter(Residente.id == x.fkresidente).first()
            persona.foto = data['foto']
            fecha = BitacoraManager(self.db).fecha_actual()
            b = Bitacora(fkusuario=data['user'], ip=data['ip'], accion="Actualizo Foto", fecha=fecha,
                         tabla="residente", identificador=persona.id)
            super().insert(b)
            self.db.merge(persona)
            self.db.commit()
            self.db.close()

            principal = self.db.query(Principal).first()
            if principal.estado == False:

                url = "http://sigas-web.herokuapp.com/api/v1/actualizar_foto"

                headers = {'Content-Type': 'application/json'}

                cadena = json.dumps(data)
                body = cadena
                resp = requests.post(url, data=body, headers=headers, verify=False)
                # response = json.loads(resp.text)

            return dict(response=None, success=True, message="Actualizado correctamente")

        except Exception as e:
            return dict(response=str(e), success=False, message="Error al actualizar")



