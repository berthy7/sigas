from .models import *
from ...common.managers import *
from sqlalchemy.exc import IntegrityError
from ...common.managers import *
from ...operaciones.bitacora.managers import *
from ...usuarios.usuario.managers import *
from ..residente.managers import *
from ..condominio.models import *
from sqlalchemy import or_
from sqlalchemy import and_
from ..modelo.models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font



class NropaseManager(SuperManager):

    def __init__(self, db):
        super().__init__(Nropase, db)

    def get_employees_tree(self):
        query = self.db.query(Condominio).filter(Condominio.estado == True).all()
        admin = dict()
        cont_tipo = 1
        for condominio in query:
            con = (condominio.id, condominio.nombre)
            admin[con] = dict()

            list_tipo_tarjeta = ['Residente', 'Visita', 'Proveedor', 'Provper', 'Excepcion']

            for tipo_tarjeta in list_tipo_tarjeta:

                tipo = (cont_tipo, tipo_tarjeta)

                cont_tipo = cont_tipo + 1

                admin[con][tipo] = dict()

                list_tarjetas = NropaseManager(self.db).listar_x_condominio_y_tipo(condominio.id,tipo_tarjeta)
                html_e = ""
                for tarjeta in list_tarjetas:


                    html = '<li class="dd-item" data-id="' + str(tarjeta.id) + str(tarjeta.id) + '"><div class="dd-handle"><input id="' + str(tarjeta.id) + str(
                        tarjeta.id) + '" data-id="' + str(tarjeta.id) + '" data-sex="' + str(tarjeta.numero) + '"type="checkbox" class="module chk-col-deep-purple employee"><label for="' + str(
                        tarjeta.id) + str(tarjeta.id) + '">' + str(tarjeta.tarjeta) + '</label></div></li>'

                    html_e = html_e + html

                    admin[con][tipo] = html_e

        return admin

    def state(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.estado = estado

        if estado:
            mensaje = "Habilito Tarjeta"
            sincro = dict(codigo=x.id, tarjeta=x.tarjeta, situacion="Acceso", fkdispositivo=id,
                          fkcondominio=x.condominios[0].fkcondominio)
        else:
            mensaje = "Deshabilito Tarjeta"
            sincro = dict(codigo=x.id, tarjeta=x.tarjeta, situacion="Denegado", fkdispositivo=id,
                          fkcondominio=x.condominios[0].fkcondominio)

        ConfiguraciondispositivoManager(self.db).funcion_configuracion_dispositivo(sincro)

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion=mensaje, fecha=fecha,
                     tabla="nropase", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        # if x.rol.nombre == "RESIDENTE":
        #
        #     resi = self.db.query(Residente).filter(Residente.id == x.fkresidente).first()
        #     resiacce = self.db.query(ResidenteAcceso).filter(ResidenteAcceso.fkresidente == x.fkresidente).first()
        #     resi.estado = estado
        #     resiacce.estado = estado
        #     resi = self.db.merge(resi)
        #     self.db.merge(resiacce)
        #     self.db.commit()
        #
        #     if x.condominio.singuardia:
        #         UsuarioManager(self.db).sincronizar_dispositivos(x, estado, resi)
        #
        # principal = self.db.query(Principal).first()
        #
        # if principal.estado:
        #
        #     try:
        #         if x.fkcondominio:
        #
        #             if x.condominio.ip_publica != "":
        #                 diccionary = dict(id=id, estado=estado, user=user, ip=ip)
        #
        #                 url = "http://" + x.condominio.ip_publica + ":" + x.condominio.puerto + "/api/v1/sincronizar_usuario_estado"
        #
        #                 headers = {'Content-Type': 'application/json'}
        #                 string = diccionary
        #                 cadena = json.dumps(string)
        #                 body = cadena
        #                 resp = requests.post(url, data=body, headers=headers, verify=False)
        #                 response = json.loads(resp.text)
        #
        #                 print(response)
        #
        #
        #     except Exception as e:
        #         # Other errors are possible, such as IOError.
        #         print("Error de conexion: " + str(e))

        return x

    def get_all(self):
        return self.db.query(self.entity).all()

    def obtener_x_tarjeta(self,tarjeta):
        return self.db.query(self.entity).filter(self.entity.tarjeta == tarjeta).first()

    def list_all(self):
        return dict(objects=self.db.query(self.entity))

    def listar_todo(self):
        return self.db.query(self.entity).order_by(self.entity.tarjeta.asc()).all()

    def listar_x_condominio(self, idcondominio):
        x = self.db.query(self.entity).join(CondominioPases).filter(CondominioPases.fkcondominio == idcondominio).all()
        return x

    def listar_x_condominio_y_tipo(self, idcondominio,tipo_tarjeta):
        x = self.db.query(self.entity).join(CondominioPases)\
            .filter(CondominioPases.fkcondominio == idcondominio)\
            .filter(self.entity.tipo == tipo_tarjeta) \
            .order_by(self.entity.tarjeta.asc()).all()
        return x

    def listar_x_tipo(self, usuario,tipopase):


        if usuario.sigas:

            if tipopase == "Proveedor" or tipopase == "Taxi":
                return self.db.query(self.entity).filter(
                    and_(self.entity.tipo != "Residente", self.entity.tipo != "Provper", self.entity.tipo != "Visita")).filter(
                    self.entity.situacion != "Ocupado").order_by(
                    self.entity.numero.asc()).all()
            else:
                return self.db.query(self.entity).filter(
                    and_(self.entity.tipo != "Residente", self.entity.tipo != "Provper", self.entity.tipo != "Proveedor")).filter(
                    self.entity.situacion != "Ocupado").order_by(
                    self.entity.numero.asc()).all()


        if usuario.rol.nombre != "RESIDENTE":

            if tipopase == "Proveedor" or tipopase == "Taxi":

                return self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(
                    Condominio.id == usuario.fkcondominio).filter(
                    and_(self.entity.tipo != "Residente", self.entity.tipo != "Provper", self.entity.tipo != "Visita")).filter(
                    self.entity.situacion != "Ocupado").order_by(Nropase.numero.asc()).all()
            else:
                return self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(
                    Condominio.id == usuario.fkcondominio).filter(
                    and_(self.entity.tipo != "Residente", self.entity.tipo != "Provper", self.entity.tipo != "Proveedor")).filter(
                    self.entity.situacion != "Ocupado").order_by(Nropase.numero.asc()).all()

        else:
            return None

    def listar_numero_pases(self,usuario):

        if usuario.sigas:
            return self.db.query(self.entity).filter(and_(self.entity.tipo != "Residente",self.entity.tipo != "Provper")).filter(self.entity.situacion != "Ocupado").order_by(
                self.entity.numero.asc()).all()

        if usuario.rol.nombre != "RESIDENTE":
            return self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(Condominio.id == usuario.fkcondominio).filter(and_(Nropase.tipo != "Residente",Nropase.tipo != "Provper")).filter(Nropase.situacion != "Ocupado").order_by(Nropase.numero.asc()).all()

        else:
            return None

    def listar_numero_pases_residente(self,usuario):

        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.tipo == "Residente").filter(self.entity.situacion != "Ocupado").order_by(
                self.entity.numero.asc()).all()

        if usuario.rol.nombre != "RESIDENTE":
            return self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(Condominio.id == usuario.fkcondominio).filter(Nropase.tipo == "Residente").filter(self.entity.situacion != "Ocupado").order_by(Nropase.numero.asc()).all()

        else:
            return None

    def listar_tarjetas_provper(self,usuario):

        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.tipo == "Provper").filter(self.entity.situacion != "Ocupado").order_by(
                self.entity.numero.asc()).all()

        if usuario.rol.nombre != "RESIDENTE":
            return self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(Condominio.id == usuario.fkcondominio).filter(Nropase.tipo == "Provper").filter(self.entity.situacion != "Ocupado").order_by(Nropase.numero.asc()).all()

        else:
            return None

    def listar_todo_condominio(self):
        list = {}
        vector = []
        c = 0

        objeto = self.db.query(self.entity).all()

        for x in objeto:

            list[c] = dict(gestion=x.fecha.year)

            c = c + 1
            vector.append(x.fecha.year)

        return list

    def insert(self, diccionary):
        if diccionary['tipo'] == "":
            diccionary['tipo'] = "Visita"

        objeto = NropaseManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Nropase.", fecha=fecha,tabla="nropase", identificador=a.id)
        super().insert(b)
        return a


    def insert_sincronizacion(self, diccionary):

        for tar in diccionary['tarjetas']:
            t = self.db.query(self.entity).filter(self.entity.id == tar['id']).first()

            if t.estado != tar['estado']:
                NropaseManager(self.db).state(tar['id'],tar['estado'],diccionary['user'],diccionary['ip'])


    def update(self, diccionary):
        objeto = NropaseManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Nropase.", fecha=fecha,tabla="nropase", identificador=a.id)
        super().insert(b)
        return a

    def delete(self, id,estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.estado = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Nropase.", fecha=fecha, tabla="nropase", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        return x

    def situacion(self, id,situacion):
        x = self.db.query(Nropase).filter(Nropase.id == id).first()
        if x.tipo != "Excepcion":
            if x:
                x.situacion = situacion

                self.db.merge(x)
                self.db.commit()

        return x

    def importar_excel(self, cname,user,ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['NUMERO_DE_PASE','TARJETA','TIPO','COD_CONDOMINIO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):

                    nropase = row[indices['NUMERO_DE_PASE']].value
                    tarjeta = row[indices['TARJETA']].value
                    tipo = row[indices['TIPO']].value
                    cod_condominio = row[indices['COD_CONDOMINIO']].value

                    print(str(nropase) + " " + str(tarjeta))

                    if tarjeta is not None:
                        if cod_condominio:
                            cod_condominio = cod_condominio.replace(" ", "")

                        list_condominio = list()

                        query = self.db.query(Nropase).filter(Nropase.tarjeta == str(tarjeta)).first()
                        query_condominio = self.db.query(Condominio).filter(
                            Condominio.codigo == str(cod_condominio)).first()

                        if not query:

                            if tipo:
                                tipo = tipo.replace(" ", "")

                            if tipo is None or tipo == "":
                                tipo = "Invitacion"


                            if nropase is None or nropase == "":
                                nropase = " "


                            if query_condominio:
                                list_condominio.append(dict(fkcondominio=query_condominio.id))

                            mode = Nropase(numero=str(nropase),tarjeta=str(tarjeta),tipo=str(tipo),condominios=list_condominio)
                            self.db.add(mode)
                            self.db.commit()

                        else:
                            query_cond = self.db.query(Nropase).join(CondominioPases).join(Condominio).filter(
                                Condominio.codigo == str(cod_condominio)).filter(
                                Nropase.tarjeta == str(tarjeta)).first()

                            if not query_cond:

                                if query_condominio:
                                    mode = CondominioPases(fknropase=query.id, fkcondominio=query_condominio.id)
                                    self.db.add(mode)
                                    self.db.commit()

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

