from .models import *
from ...common.managers import *
from sqlalchemy.exc import IntegrityError
from ...common.managers import *
from ...operaciones.bitacora.managers import *
from ...usuarios.usuario.managers import *
from ..residente.managers import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side



class InvitadoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Invitado, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.estado == True).order_by(self.entity.apellidop.asc()).all()

    def listar_x_residente(self,usuario):

        if usuario.sigas:
            return self.db.query(self.entity).filter(self.entity.estado == True).order_by(
                self.entity.apellidop.asc()).all()
        elif usuario.rol.nombre == "RESIDENTE":
            return self.db.query(self.entity).join(Amistad).join(Residente).filter(self.entity.estado == True).filter(
                Residente.id == usuario.fkresidente).order_by(self.entity.apellidop.asc()).all()
        else:
            return self.db.query(self.entity).filter(self.entity.estado == True).order_by(
                self.entity.apellidop.asc()).all()


    def obtener_x_id(self,idinvitado):
        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.id == idinvitado).first()

    def obtener_x_ci(self,ciinvitado):

        return self.db.query(self.entity).filter(self.entity.estado == True).filter(self.entity.ci == ciinvitado).first()

    def insert(self, diccionary):
        usuario = UsuarioManager(self.db).get_by_pass(diccionary['user'])

        invitado = InvitadoManager(self.db).obtener_x_ci(diccionary['ci'])

        if invitado is None:
            objeto = InvitadoManager(self.db).entity(**diccionary)
            fecha = BitacoraManager(self.db).fecha_actual()
            objeto.vehiculos = []

            a = super().insert(objeto)
            print("registro contacto: " + str(a.id))
            b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Invitado.", fecha=fecha,tabla="invitado", identificador=a.id)
            super().insert(b)
        else:
            if invitado.nombre == "":
                invitado.nombre = diccionary['nombre']
            if invitado.apellidop == "":
                invitado.apellidop = diccionary['apellidop']
            if invitado.apellidom == "":
                invitado.apellidom = diccionary['apellidom']
            if invitado.sexo is None:
                invitado.sexo = diccionary['sexo']
            if invitado.expendido is None:
                invitado.expendido = diccionary['expendido']
            if invitado.telefono is None:
                invitado.telefono = diccionary['telefono']
            a = super().update(invitado)

        if usuario.fkresidente != None:
            amistad = Amistad(fkresidente=usuario.fkresidente, fkinvitado=a.id)
            super().insert(amistad)

        for vehi in diccionary['vehiculos']:
            VehiculoManager(self.db).registrar_vehiculo_invitado(vehi,a.id)

        return a

    def update(self, diccionary):

        objeto = InvitadoManager(self.db).entity(**diccionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Invitado.", fecha=fecha,tabla="invitado", identificador=a.id)
        super().insert(b)

        return a


    def registrar_invitado(self, diccionario):
        usuario = UsuarioManager(self.db).get_by_pass(diccionario['user'])
        nombre = diccionario['nombre']
        apellidop = diccionario['apellidop']
        apellidom = diccionario['apellidom']
        ci = diccionario['ci']
        expendido = diccionario['expendido']

        invitado = InvitadoManager(self.db).obtener_x_ci(ci)

        if invitado:
            return invitado
        else:
            dict_invitado = dict(nombre=nombre,apellidop=apellidop,apellidom=apellidom,ci=ci, expendido=expendido)

            objeto = InvitadoManager(self.db).entity(**dict_invitado)
            fecha = BitacoraManager(self.db).fecha_actual()
            a = super().insert(objeto)
            b = Bitacora(fkusuario=diccionario['user'], ip=diccionario['ip'], accion="Registro Invitado.", fecha=fecha,
                         tabla="invitado", identificador=a.id)
            super().insert(b)

        if usuario.fkresidente != None:
            amistad = Amistad(fkresidente=usuario.fkresidente, fkinvitado=a.id)
            super().insert(amistad)

        return a

    def registrar_conductor(self, diccionario):
        usuario = UsuarioManager(self.db).get_by_pass(diccionario['user'])
        nombre = diccionario['nombre_conductor']
        apellidop = diccionario['apellidop_conductor']
        apellidom = diccionario['apellidom_conductor']
        ci = diccionario['ci_conductor']
        expendido = diccionario['expendido_conductor']

        invitado = InvitadoManager(self.db).obtener_x_ci(ci)

        if invitado:
            return invitado
        else:
            dict_invitado = dict(nombre=nombre, apellidop=apellidop, apellidom=apellidom, ci=ci, expendido=expendido)

            objeto = InvitadoManager(self.db).entity(**dict_invitado)
            fecha = BitacoraManager(self.db).fecha_actual()
            a = super().insert(objeto)
            b = Bitacora(fkusuario=diccionario['user'], ip=diccionario['ip'], accion="Registro conductor de taxi.", fecha=fecha,
                         tabla="invitado", identificador=a.id)
            super().insert(b)

        return a

    def actualizar_invitado(self, diccionario):

        id = diccionario['fkinvitado']
        nombre = diccionario['nombre']
        apellidop = diccionario['apellidop']
        apellidom = diccionario['apellidom']
        ci = diccionario['ci']
        expendido = diccionario['expendido']


        invitado = InvitadoManager(self.db).obtener_x_id(id)

        invitado.nombre = nombre
        invitado.apellidop = apellidop
        invitado.apellidom = apellidom
        invitado.ci = ci
        invitado.expendido = expendido


        super().update(invitado)

    def actualizar_conductor(self, diccionario):

        id = diccionario['fkconductor']
        nombre = diccionario['nombre_conductor']
        apellidop = diccionario['apellidop_conductor']
        apellidom = diccionario['apellidom_conductor']
        ci = diccionario['ci_conductor']
        expendido = diccionario['expendido_conductor']

        invitado = InvitadoManager(self.db).obtener_x_id(id)

        invitado.nombre = nombre
        invitado.apellidop = apellidop
        invitado.apellidom = apellidom
        invitado.ci = ci
        invitado.expendido = expendido

        super().update(invitado)


    def invitado_excel(self,):
        cname = "Invitados.xlsx"

        invitados = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'NOMBRE'
        ws['C' + str(indice)] = 'APELLIDOP'
        ws['D' + str(indice)] = 'APELLIDOM'
        ws['E' + str(indice)] = 'SEXO'
        ws['F' + str(indice)] = 'CI'
        ws['G' + str(indice)] = 'EXPEDIDO'
        ws['H' + str(indice)] = 'TELEFONO'
        ws['I' + str(indice)] = 'DESCRIPCION'
        ws['J' + str(indice)] = 'PERMANENTE'
        ws['K' + str(indice)] = 'FKNROPASE'
        ws['L' + str(indice)] = 'ESTADO'

        for i in invitados:

            indice = indice + 1
            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.nombre
            ws['C' + str(indice)] = i.apellidop
            ws['D' + str(indice)] = i.apellidom
            ws['E' + str(indice)] = i.sexo
            ws['F' + str(indice)] = i.ci
            ws['G' + str(indice)] = i.expendido
            ws['H' + str(indice)] = i.telefono
            ws['I' + str(indice)] = i.descripcion
            ws['J' + str(indice)] = i.permanente
            ws['K' + str(indice)] = i.fknropase
            ws['L' + str(indice)] = i.estado


        wb.save("server/common/resources/downloads/" + cname)
        return cname


    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID','NOMBRE', 'APELLIDOP', 'APELLIDOM', 'SEXO','CI', 'EXPEDIDO', 'TELEFONO', 'DESCRIPCION', 'PERMANENTE', 'FKNROPASE', 'ESTADO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    id = row[indices['ID']].value
                    nombre = row[indices['NOMBRE']].value
                    apellidop = row[indices['APELLIDOP']].value
                    apellidom = row[indices['APELLIDOM']].value
                    sexo = row[indices['SEXO']].value
                    ci = row[indices['CI']].value
                    expedido = row[indices['EXPEDIDO']].value
                    telefono = row[indices['TELEFONO']].value
                    descripcion = row[indices['DESCRIPCION']].value
                    permanente = row[indices['PERMANENTE']].value
                    fknropase = row[indices['FKNROPASE']].value
                    estado = row[indices['ESTADO']].value

                    if id is not None:

                        query = self.db.query(Invitado).filter(Invitado.id == id).first()

                        if not query:



                            invi = Invitado(id=id,nombre=nombre,apellidop=apellidop,apellidom=apellidom,
                                            sexo=sexo, ci=ci, expedido=expedido,telefono=telefono,
                                            descripcion=descripcion, permanente=permanente, fknropase=fknropase, estado=estado)

                            self.db.merge(invi)
                            self.db.flush()

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}


